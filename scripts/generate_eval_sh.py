#!/usr/bin/env python3
"""
读取 test_run_sh.xlsx 第1列(groundtruth)和第2列(推理结果)，
融合数据后直接调用 app.services.evaluate_service.run_evaluation （即 POST
/api/v1/evaluate/case 后端逻辑）进行评测，结果写入第4列。第3列保存请求
body JSON 以便排查。

多轮 + 合并扩展：
- ``--rounds N``：把同一份 case 评测 N 次。body 仍只写一列（同一份 JSON），
  N 个 result 写入连续 N 列，header 加 ``_rK`` 后缀（``eval_result_r1`` …）。
  ``rounds=1`` 时退化为旧行为：header 仍是 ``eval_result``，向后兼容。
- ``--merge``：要求 ``--rounds >= 3``。在 N 轮 result 之后再追加两列：
  ``merge_results`` 是前 3 轮 IISR/IFS 多数表决合并后的完整 eval_result JSON；
  ``iisr_merge_notes`` 是每条 rubric 的合并备注（"3 个都相同" / "有 2 个相同"
  / "3 个都不同"）。合并逻辑复用 ``scripts/merge_iisr_ifs_majority.py``。

用法：
    python scripts/generate_eval_sh.py [xlsx_file]
    python scripts/generate_eval_sh.py data/outputs/simulator_res/test_run_sh.xlsx
    python scripts/generate_eval_sh.py data/outputs/simulator_res/test_run_sh.xlsx --rounds 3 --merge
"""

import argparse
import asyncio
import json
import sys
import traceback
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 让脚本能从仓库根 import app.* 模块
ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


def clean_json_str(raw: str) -> str:
    """清理 Excel 中的非标准空格(\xa0)等字符，使 JSON 可解析。"""
    if not isinstance(raw, str):
        return ""
    return raw.replace("\xa0", " ").strip()


def parse_json_safe(raw: str) -> dict | None:
    """安全解析 JSON 字符串。

    用 ``raw_decode`` 而非 ``loads``：上游 xlsx 里部分单元格末尾会带一个
    多余的逗号（``},``）或重复对象，``loads`` 会报 ``Extra data``，但
    ``raw_decode`` 只返回第一个完整 JSON 对象，把尾部杂质忽略掉。
    """
    cleaned = clean_json_str(raw)
    if not cleaned:
        return None
    try:
        decoder = json.JSONDecoder()
        obj, _end = decoder.raw_decode(cleaned)
        return obj
    except json.JSONDecodeError as e:
        # 真的格式坏到第一个对象都解析不出来：再试一次，剥掉首尾常见的脏字符
        stripped = cleaned.lstrip(",;\n\r\t ").rstrip(",;\n\r\t ")
        try:
            decoder = json.JSONDecoder()
            obj, _end = decoder.raw_decode(stripped)
            return obj
        except json.JSONDecodeError:
            print(f"  ⚠️  JSON 解析失败: {e}")
            return None


def extract_fields_from_ground_truth(gt_data: dict) -> dict:
    """从第1列的 groundtruth JSON 中提取所需字段。"""
    input_data = gt_data.get("input_data", {})
    user_simulator = gt_data.get("user_simulator", {})
    context = input_data.get("context", {})

    query = user_simulator.get("query") or input_data.get("query", "")
    full_intent = user_simulator.get("full_intent", "")
    current_time = context.get("time", "")
    current_location = context.get("user_loc_name", "")

    return {
        "task_id": gt_data.get("task_id", ""),
        "query": query,
        "full_intent": full_intent,
        "ground_truth": gt_data.get("ground_truth", {}),
        "current_time": current_time,
        "current_location": current_location,
    }


def build_merged_case(
    extracted_gt: dict,
    result_data: dict,
    *,
    enable_verification: bool = False,
    enable_meta_judge: bool = False,
) -> dict:
    """融合第1列和第2列的数据，生成评测 case JSON（即 EvaluateRequest body）。

    ``enable_verification`` 控制是否调 GoogleWebSearch 核验 IFS 标记
    ``need_external_verify=true`` 的要素 + 老 ``fact_summary`` 的
    ``need_verify=true`` 事实；默认关闭以保持脚本原有"纯离线"行为。
    """
    conversation_messages = result_data.get("conversation_history_messages", [])
    session_stats = result_data.get("session_stats", {}) or {}

    return {
        "case_id": extracted_gt["task_id"],
        "query": extracted_gt["query"],
        "full_intent": extracted_gt["full_intent"],
        "persona": "",
        "current_time": extracted_gt["current_time"],
        "current_location": extracted_gt["current_location"],
        "language": "chinese",
        "enable_verification": enable_verification,
        "enable_meta_judge": enable_meta_judge,
        "ground_truth": extracted_gt["ground_truth"],
        "session_stats": session_stats,
        "conversation_history_messages": conversation_messages,
    }


async def evaluate_case_in_process(merged_case: dict) -> dict:
    """直接调用 evaluate_service.run_evaluation，绕过 HTTP 层。"""
    # 延迟导入：避免无关命令（--help）也加载重型依赖
    from app.schemas.evaluate_schemas import EvaluateRequest
    from app.services.evaluate_service import run_evaluation

    req = EvaluateRequest.model_validate(merged_case)
    resp = await run_evaluation(req)
    return resp.model_dump()


async def run_all_cases(
    cases: list[tuple[int, dict]], concurrency: int
) -> list[tuple[int, str, bool]]:
    """并发执行评测，返回 (excel_row, cell_text, ok) 列表。

    - excel_row:  对应 Excel 行号（写第4列时用）
    - cell_text:  写入第4列的内容（成功时为 JSON，失败时为 traceback）
    - ok:         是否成功
    """
    sem = asyncio.Semaphore(concurrency)

    async def one(excel_row: int, case: dict) -> tuple[int, str, bool]:
        async with sem:
            print(f"[row {excel_row}] start")
            try:
                result = await evaluate_case_in_process(case)
                text = json.dumps(result, ensure_ascii=False, indent=2)
                print(f"[row {excel_row}] ✓ done")
                return excel_row, text, True
            except Exception as e:
                tb = traceback.format_exc()
                print(f"[row {excel_row}] ✗ {e}")
                return excel_row, f"ERROR: {e}\n\n{tb}", False

    return await asyncio.gather(*(one(r, c) for r, c in cases))


def extract_ar_details(eval_result: dict) -> dict | None:
    """从评测结果中提取 results.details.AR 模块。"""
    if not isinstance(eval_result, dict):
        return None
    results = eval_result.get("results")
    if not isinstance(results, dict):
        return None
    details = results.get("details")
    if not isinstance(details, dict):
        return None
    ar = details.get("AR")
    if isinstance(ar, dict):
        return ar
    return None


def extract_ifs_details(eval_result: dict) -> dict | None:
    """从评测结果中提取 results.details.IFS 模块。"""
    if not isinstance(eval_result, dict):
        return None
    results = eval_result.get("results")
    if not isinstance(results, dict):
        return None
    details = results.get("details")
    if not isinstance(details, dict):
        return None
    ifs = details.get("IFS")
    if isinstance(ifs, dict):
        return ifs
    return None


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "xlsx_file",
        nargs="?",
        default="data/outputs/simulator_res/test_run_sh.xlsx",
        help="输入 xlsx 文件路径（默认: data/outputs/simulator_res/test_run_sh.xlsx）",
    )
    parser.add_argument(
        "--no-execute", action="store_true",
        help="只生成第3列 body JSON，不执行评测（不写第4列）",
    )
    parser.add_argument(
        "--concurrency", type=int, default=5,
        help="并发执行的评测任务数（默认 5）。受 LLM 端速率限制约束，太大反而会被限流。",
    )
    parser.add_argument(
        "--enable-verification", action="store_true",
        help=(
            "打开外部事实核验（FactVerifier → GoogleWebSearch via MCP）。"
            "默认关闭以保持纯离线行为；打开后 IFS 标 need_external_verify=true "
            "的要素 + fact_summary 标 need_verify=true 的事实会走 web 搜索 + LLM 判定。"
        ),
    )
    parser.add_argument(
        "--enable-meta-judge", action="store_true",
        help="打开 Stage 4.5 Meta-Judge 审计（Devil's Advocate）。默认关闭。",
    )
    parser.add_argument(
        "--extract-ar-ifs", action="store_true",
        help="从第4列评测结果中提取 AR 和 IFS 模块，分别写入第5列和第6列",
    )
    parser.add_argument(
        "--task-id", default=None,
        help=(
            "只跑第1列中包含该 task_id 子串的行；命中行先复制到 "
            "<原文件名>_filtered_<task_id 前8位>.xlsx，再在副本上跑评测，"
            "**不会动原文件**。"
        ),
    )
    parser.add_argument(
        "--rounds", type=int, default=1,
        help="评测轮数。>1 时同一份 body 跑 N 次，结果依次写入第 4..(3+N) 列，"
             "header 加 _rK 后缀（eval_result_r1 ...）。默认 1（向后兼容）。",
    )
    parser.add_argument(
        "--merge", action="store_true",
        help="要求 --rounds>=3。在 N 轮 result 之后追加 2 列："
             "merge_results 是前 3 轮 IISR/IFS 多数表决合并后的完整 eval_result JSON；"
             "iisr_merge_notes 是每条 rubric 的合并备注。",
    )
    args = parser.parse_args()

    if args.rounds < 1:
        print(f"--rounds 必须 >= 1，收到 {args.rounds}")
        sys.exit(1)
    if args.merge and args.rounds < 3:
        print(f"--merge 要求 --rounds >= 3，当前 rounds={args.rounds}")
        sys.exit(1)
    if args.extract_ar_ifs and args.rounds > 1:
        print("⚠️  --extract-ar-ifs 在 --rounds>1 时仅作用于第 1 轮结果（第4列）")

    xlsx_path = Path(args.xlsx_file)
    if not xlsx_path.is_absolute():
        xlsx_path = ROOT_DIR / xlsx_path
    if not xlsx_path.exists():
        print(f"文件不存在: {xlsx_path}")
        sys.exit(1)

    # --task-id：把命中行另存到副本，后续都在副本上操作，保护原文件
    if args.task_id:
        df_full = pd.read_excel(xlsx_path)
        mask = df_full.iloc[:, 0].astype(str).str.contains(args.task_id, na=False)
        n_hit = int(mask.sum())
        if n_hit == 0:
            print(f"task_id={args.task_id!r} 在第1列中未匹配到任何行")
            sys.exit(1)
        filtered = xlsx_path.with_name(
            f"{xlsx_path.stem}_filtered_{args.task_id[:8]}.xlsx"
        )
        df_full[mask].to_excel(filtered, index=False)
        print(f"[filter] task_id={args.task_id} 命中 {n_hit} 行 → {filtered.name}")
        xlsx_path = filtered

    print(f"读取文件: {xlsx_path}")
    df = pd.read_excel(xlsx_path)
    print(f"共 {len(df)} 行数据\n")

    wb = load_workbook(xlsx_path)
    ws = wb.active

    if ws.cell(row=1, column=3).value is None:
        ws.cell(row=1, column=3, value="eval_request_body")

    # 第 4..(3+N) 列分别是各轮 result。rounds=1 时 header 保持 "eval_result"
    # 以向后兼容旧消费者；rounds>1 时全部加 _rK 后缀。
    if args.rounds == 1:
        result_headers = ["eval_result"]
    else:
        result_headers = [f"eval_result_r{k + 1}" for k in range(args.rounds)]
    for k, header in enumerate(result_headers):
        col = 4 + k
        if ws.cell(row=1, column=col).value is None:
            ws.cell(row=1, column=col, value=header)

    success_count = 0
    pending: list[tuple[int, dict]] = []  # (excel_row, merged_case)

    # ── Phase 1: 解析所有行，写第3列 body JSON ───────────────────────
    #
    # 两种输入形态自动识别（按第 1 列首条非空行的 JSON keys 判定）：
    #
    # * **body-mode**：第 1 列已经是完整 EvaluateRequest body（含
    #   ``case_id`` + ``conversation_history_messages`` 等键），不再读
    #   第 2 列、不再走 ``extract_fields_from_ground_truth`` +
    #   ``build_merged_case``，直接拿来用。仍把同一份 body 抄到第 3 列
    #   保持列布局不变（第 4 列起仍是 eval_result），下游 --merge /
    #   --extract-ar-ifs 全部沿用旧列号。
    #
    # * **legacy gt+result mode**：第 1 列是 groundtruth、第 2 列是推理
    #   结果，按老逻辑融合后写第 3 列 body。
    #
    # ``enable_verification`` / ``enable_meta_judge`` 在 body-mode 下用
    # CLI 参数覆盖 body 里的同名字段，保持与 legacy 一致的语义。
    nonempty = df.iloc[:, 0].dropna()
    first_row_sample = parse_json_safe(str(nonempty.iloc[0])) if len(nonempty) else None
    body_mode = (
        isinstance(first_row_sample, dict)
        and "conversation_history_messages" in first_row_sample
        and "case_id" in first_row_sample
    )
    n_cols = df.shape[1]
    print(f"[mode] {'body-mode (col1 已是 EvaluateRequest body)' if body_mode else 'legacy gt+result mode'}; df 列数={n_cols}")

    for row_idx in range(len(df)):
        excel_row = row_idx + 2
        col1_raw = str(df.iloc[row_idx, 0]) if pd.notna(df.iloc[row_idx, 0]) else ""
        if not col1_raw:
            print(f"第 {excel_row} 行跳过：第1列为空")
            continue

        if body_mode:
            merged_case = parse_json_safe(col1_raw)
            if not isinstance(merged_case, dict):
                print(f"第 {excel_row} 行跳过：第1列 body JSON 解析失败")
                continue
            # CLI 参数覆盖 body 里同名字段（默认 False）
            if args.enable_verification:
                merged_case["enable_verification"] = True
            if args.enable_meta_judge:
                merged_case["enable_meta_judge"] = True
        else:
            if n_cols < 2:
                print(f"第 {excel_row} 行跳过：legacy 模式需要第2列推理结果，但表只有 {n_cols} 列")
                continue
            result_raw = str(df.iloc[row_idx, 1]) if pd.notna(df.iloc[row_idx, 1]) else ""
            if not result_raw:
                print(f"第 {excel_row} 行跳过：第2列为空")
                continue
            gt_data = parse_json_safe(col1_raw)
            result_data = parse_json_safe(result_raw)
            if not gt_data or not result_data:
                print(f"第 {excel_row} 行跳过：JSON 解析失败")
                continue
            extracted_gt = extract_fields_from_ground_truth(gt_data)
            merged_case = build_merged_case(
                extracted_gt,
                result_data,
                enable_verification=args.enable_verification,
                enable_meta_judge=args.enable_meta_judge,
            )

        body_json = json.dumps(merged_case, ensure_ascii=False, indent=2)
        cell3 = ws.cell(row=excel_row, column=3, value=body_json)
        cell3.alignment = Alignment(wrap_text=True, vertical="top")
        success_count += 1
        pending.append((excel_row, merged_case))

    # ── Phase 2: 并发评测，每轮写一列 ────────────────────────────────
    execute_count_per_round: list[int] = []
    if not args.no_execute and pending:
        for round_idx in range(args.rounds):
            result_col = 4 + round_idx
            print(
                f"\n开始第 {round_idx + 1}/{args.rounds} 轮评测："
                f"共 {len(pending)} 行，concurrency={args.concurrency}，"
                f"写入第 {result_col} 列\n"
            )
            results = asyncio.run(run_all_cases(pending, args.concurrency))
            ok_count = 0
            for excel_row, text, ok in results:
                cell = ws.cell(row=excel_row, column=result_col, value=text)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if ok:
                    ok_count += 1
            execute_count_per_round.append(ok_count)
            # 每轮跑完都存一次，避免长跑中途中断丢失数据
            wb.save(xlsx_path)

    ws.column_dimensions["C"].width = 80
    for k in range(args.rounds):
        col_letter = ws.cell(row=1, column=4 + k).column_letter
        ws.column_dimensions[col_letter].width = 80

    # ── Phase 3: --merge，对前 3 轮 result 做 IISR/IFS 多数表决合并 ────
    merge_processed = 0
    merge_skipped = 0
    next_col = 4 + args.rounds  # 下一可用列号
    if args.merge and not args.no_execute:
        # 延迟导入：merge 脚本与本脚本在同一目录
        sys.path.insert(0, str(Path(__file__).parent))
        from merge_iisr_ifs_majority import build_merged_eval

        merge_col = next_col
        notes_col = next_col + 1
        ws.cell(row=1, column=merge_col, value="merge_results")
        ws.cell(row=1, column=notes_col, value="iisr_merge_notes")

        for row_idx in range(2, ws.max_row + 1):
            cells = [ws.cell(row=row_idx, column=4 + k).value for k in range(3)]
            if not all(cells):
                merge_skipped += 1
                continue
            parsed = [parse_json_safe(str(v or "")) for v in cells]
            if not all(isinstance(d, dict) for d in parsed):
                merge_skipped += 1
                continue
            merged, notes = build_merged_eval(parsed)
            mc = ws.cell(
                row=row_idx, column=merge_col,
                value=json.dumps(merged, ensure_ascii=False, indent=2),
            )
            mc.alignment = Alignment(wrap_text=True, vertical="top")
            nc = ws.cell(row=row_idx, column=notes_col, value="; ".join(notes))
            nc.alignment = Alignment(wrap_text=True, vertical="top")
            merge_processed += 1

        ws.column_dimensions[ws.cell(row=1, column=merge_col).column_letter].width = 80
        ws.column_dimensions[ws.cell(row=1, column=notes_col).column_letter].width = 60
        next_col += 2
        print(
            f"   merge 完成：processed={merge_processed} skipped={merge_skipped}"
            f"  merge_col=col{merge_col}  notes_col=col{notes_col}"
        )

    # ── Phase 4: --extract-ar-ifs，从第4列（第1轮 result）提取 AR/IFS ───
    if args.extract_ar_ifs:
        ar_col = next_col
        ifs_col = next_col + 1
        ws.cell(row=1, column=ar_col, value="AR_details")
        ws.cell(row=1, column=ifs_col, value="IFS_details")

        ar_count = 0
        ifs_count = 0
        for row_idx in range(2, ws.max_row + 1):
            col4_raw = ws.cell(row=row_idx, column=4).value
            if not col4_raw:
                continue

            try:
                eval_result = json.loads(clean_json_str(str(col4_raw)))
            except json.JSONDecodeError:
                continue

            ar = extract_ar_details(eval_result)
            if ar is not None:
                cell_ar = ws.cell(
                    row=row_idx, column=ar_col,
                    value=json.dumps(ar, ensure_ascii=False, indent=2),
                )
                cell_ar.alignment = Alignment(wrap_text=True, vertical="top")
                ar_count += 1

            ifs = extract_ifs_details(eval_result)
            if ifs is not None:
                cell_ifs = ws.cell(
                    row=row_idx, column=ifs_col,
                    value=json.dumps(ifs, ensure_ascii=False, indent=2),
                )
                cell_ifs.alignment = Alignment(wrap_text=True, vertical="top")
                ifs_count += 1

        ws.column_dimensions[ws.cell(row=1, column=ar_col).column_letter].width = 80
        ws.column_dimensions[ws.cell(row=1, column=ifs_col).column_letter].width = 80
        print(f"   提取 AR: {ar_count} 行，IFS: {ifs_count} 行")

    wb.save(xlsx_path)
    print(f"\n 完成！共处理 {success_count} 行，body JSON 写入第3列")
    if not args.no_execute:
        for k, ok in enumerate(execute_count_per_round):
            print(f"   第 {k + 1} 轮评测成功 {ok} 行，结果写入第 {4 + k} 列")
        if args.merge:
            print(f"   merge_results 写入第 {4 + args.rounds} 列（processed={merge_processed}）")
    print(f"   文件: {xlsx_path}")


if __name__ == "__main__":
    main()
