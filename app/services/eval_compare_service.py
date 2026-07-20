"""Eval result multi-file comparison service.

接受 N (>=2) 个 evaluation_result xlsx，复用 ``eval_summary_service`` 的 loader
分别得到每个文件的 ``BatchReport``，再做并列对比可视化。

入口 ``compare()`` 是 HTTP 路由 / CLI 共用的；CLI 层只做 argparse + 转发。
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.evaluation.metrics_summary import BatchReport, aggregate_batch
from app.services.eval_summary_service import (
    EvalRow,
    METRIC_NAMES,
    RunMeans,
    StatusCounts,
    VisualizeMode,
    build_metric_scores,
    compute_run_means,
    load_eval_results,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------


@dataclass
class RunSnapshot:
    """单个 xlsx 运行的快照。"""

    label: str
    xlsx_path: Path
    report: BatchReport
    rows: list[EvalRow]
    counts: StatusCounts


@dataclass
class ComparisonProducts:
    """compare() 的产物落盘路径。

    主报告为 ``report.md``（纯文本，无图嵌入）；
    ``comparison_summary.json`` 给机器消费；
    ``comparison_summary.xlsx`` 给数据透视；
    ``charts/*.png`` 是可选的独立图表文件，不在 md 里引用。
    """

    out_dir: Path
    json_path: Path
    md_path: Path
    xlsx_path: Optional[Path] = None
    charts_dir: Optional[Path] = None
    chart_paths: dict[str, Path] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for c in col:
            if c.value is not None:
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ---------------------------------------------------------------------------
# Loader / orchestration
# ---------------------------------------------------------------------------


def _load_one(label: str, path: Path) -> RunSnapshot:
    """对单个 xlsx 跑一次 load + aggregate。

    7 指标 + CEI 分母 = 全部行（含失败 case，按 0 计入）；runtime（E2E_Latency
    / tokens）只取 success 行，避免物理量被 0 拉低均值 / 污染 p95。
    """
    rows, counts = load_eval_results(path)
    scores = build_metric_scores(rows)
    success_rows = [r for r in rows if r.status == "success"]
    ttft = [r.e2e_latency_ms for r in success_rows]
    tokens = [r.total_tokens for r in success_rows]
    cei = [float(r.metrics.get("CEI", 0.0) or 0.0) for r in rows]
    report = aggregate_batch(
        scores,
        e2e_latency_ms_per_case=ttft,
        tokens_per_case=tokens,
        cei_per_case=cei,
    )
    return RunSnapshot(
        label=label, xlsx_path=path, report=report, rows=rows, counts=counts
    )


def compare(
    xlsx_paths: list[Path],
    output_dir: Path,
    *,
    labels: list[str] | None = None,
    visualize: VisualizeMode = "both",
) -> ComparisonProducts:
    """对 N (>=2) 个 evaluation_result xlsx 做并列对比。

    Args:
        xlsx_paths: 至少 2 个；少于 2 抛 ValueError
        output_dir: 落盘根目录；产物会落在 ``<output_dir>/compare/<ts>/``
        labels: 显示用名称；None 时取每个 xlsx 的 ``Path.stem``
        visualize: ``xlsx`` / ``png`` / ``both`` / ``none``

    Returns:
        ComparisonProducts 含所有产物路径
    """
    if len(xlsx_paths) < 2:
        raise ValueError(
            "compare 至少需要 2 个 xlsx 文件；单文件请用 aggregate_eval_results"
        )
    if labels is None:
        labels = [p.stem for p in xlsx_paths]
    if len(labels) != len(xlsx_paths):
        raise ValueError(
            f"labels 数量 ({len(labels)}) 与 xlsx_paths ({len(xlsx_paths)}) 不一致"
        )

    snapshots = [_load_one(lab, p) for lab, p in zip(labels, xlsx_paths)]

    for s in snapshots:
        logger.info(
            f"[{s.label}] {s.xlsx_path.name}: total={s.counts.n_total} "
            f"success={s.counts.n_success} skipped={s.counts.n_skipped} "
            f"error={s.counts.n_error} unparseable={s.counts.n_unparseable}"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    # 与单文件统计 (report/single/<ts>/) 对称：多文件对比走 report/compare/<ts>/，
    # 让两类报告在同一父目录下分组隔离、便于管理
    base_dir = output_dir / "compare" / ts
    base_dir.mkdir(parents=True, exist_ok=True)

    json_path = base_dir / "comparison_summary.json"
    md_path = base_dir / "report.md"
    render_json(snapshots, json_path)
    render_md(snapshots, md_path)

    products = ComparisonProducts(
        out_dir=base_dir, json_path=json_path, md_path=md_path
    )

    if visualize in ("xlsx", "both"):
        xlsx_out = base_dir / "comparison_summary.xlsx"
        render_xlsx(snapshots, xlsx_out)
        products.xlsx_path = xlsx_out

    if visualize in ("png", "both"):
        charts_dir = base_dir / "charts"
        chart_paths = render_png(snapshots, charts_dir)
        products.charts_dir = charts_dir if chart_paths else None
        products.chart_paths = chart_paths

    # 控制台直接打印 md 内容（人眼可读）
    print(md_path.read_text(encoding="utf-8"))
    return products


# ---------------------------------------------------------------------------
# Renderers — Markdown / JSON
# ---------------------------------------------------------------------------


def _md_metric_summary_table(snapshots: list[RunSnapshot]) -> str:
    """顶部「按运行汇总」表：每个 xlsx 一行，列出 6 指标 mean + runtime mean。"""
    headers = (
        ["Label", "n_total", "n_success"]
        + [f"{m}.mean" for m in METRIC_NAMES]
        + ["E2E_Latency mean (ms)", "Tokens mean"]
    )
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for s in snapshots:
        cells: list[str] = [
            s.label,
            str(s.counts.n_total),
            str(s.counts.n_success),
        ]
        for m in METRIC_NAMES:
            ms = getattr(s.report, m.lower())
            cells.append(f"{ms.mean:.4f}")
        cells.append(
            f"{s.report.e2e_latency_ms.mean:.0f}" if s.report.e2e_latency_ms else "—"
        )
        cells.append(
            f"{s.report.tokens_per_task.mean:.0f}"
            if s.report.tokens_per_task
            else "—"
        )
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def _md_detail_block(s: RunSnapshot) -> str:
    """单个运行的详细块：6 指标完整 stats + runtime。"""
    lines: list[str] = []
    lines.append(f"### {s.label} — `{s.xlsx_path.name}`")
    lines.append("")
    lines.append(
        f"样本：total={s.counts.n_total} · success={s.counts.n_success} · "
        f"skipped={s.counts.n_skipped} · error={s.counts.n_error} · "
        f"unparseable={s.counts.n_unparseable}"
    )
    lines.append("")
    lines.append("| metric | n | mean | median | std | min | max |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- |")
    for m in METRIC_NAMES:
        ms = getattr(s.report, m.lower())
        lines.append(
            f"| {ms.metric} | {ms.n} | {ms.mean:.4f} | {ms.median:.4f} | "
            f"{ms.std:.4f} | {ms.min:.4f} | {ms.max:.4f} |"
        )
    lines.append("")

    rt_lines: list[str] = []
    if s.report.e2e_latency_ms:
        r = s.report.e2e_latency_ms
        rt_lines.append(
            f"- **E2E_Latency (ms)**：N={r.count} · mean={r.mean:.0f} · "
            f"p50={r.p50:.0f} · p95={r.p95:.0f} · p99={r.p99:.0f} · "
            f"min={r.min:.0f} · max={r.max:.0f}"
        )
    if s.report.tokens_per_task:
        r = s.report.tokens_per_task
        rt_lines.append(
            f"- **Tokens / task**：N={r.count} · mean={r.mean:.0f} · "
            f"p50={r.p50:.0f} · p95={r.p95:.0f} · "
            f"min={r.min:.0f} · max={r.max:.0f} · sum={r.sum:.0f}"
        )
    if rt_lines:
        lines.append("Runtime:")
        lines.extend(rt_lines)
        lines.append("")
    return "\n".join(lines)


def render_md(snapshots: list[RunSnapshot], path: Path) -> Path:
    """主报告：顶部汇总表 + 每个 xlsx 详细块。纯文本，无图嵌入。"""
    n = len(snapshots)
    lines: list[str] = []
    lines.append("# 评测对比报告")
    lines.append("")
    lines.append(f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- 评测数：{n}")
    lines.append("")

    # 第一部分：汇总
    lines.append("## 一、汇总（按运行）")
    lines.append("")
    lines.append(_md_metric_summary_table(snapshots))
    lines.append("")

    # 第二部分：详细结果（每个 xlsx 一段）
    lines.append("## 二、详细结果")
    lines.append("")
    for s in snapshots:
        lines.append(_md_detail_block(s))

    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return path


def render_json(snapshots: list[RunSnapshot], path: Path) -> Path:
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "n_runs": len(snapshots),
        "runs": [
            {
                "label": s.label,
                "xlsx_path": str(s.xlsx_path),
                "counts": {
                    "n_total": s.counts.n_total,
                    "n_success": s.counts.n_success,
                    "n_skipped": s.counts.n_skipped,
                    "n_error": s.counts.n_error,
                    "n_unparseable": s.counts.n_unparseable,
                },
                "report": s.report.as_dict(),
            }
            for s in snapshots
        ],
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return path


# ---------------------------------------------------------------------------
# Renderer — XLSX (option A)
# ---------------------------------------------------------------------------


def render_xlsx(snapshots: list[RunSnapshot], path: Path) -> Path:
    """生成 3-sheet 对比 xlsx：Summary / Metrics / Runtime + 嵌入 grouped BarChart。"""
    n = len(snapshots)
    labels = [s.label for s in snapshots]

    wb = Workbook()

    # ─── Sheet 1: Summary ───
    ws = wb.active
    ws.title = "Summary"

    ws.append(["count"] + labels)
    _style_header(ws, 1)
    for fname in ("n_total", "n_success", "n_skipped", "n_error", "n_unparseable"):
        ws.append([fname] + [getattr(s.counts, fname) for s in snapshots])

    ws.append([])
    metric_header_row = ws.max_row + 1
    ws.append(["metric"] + labels)
    _style_header(ws, metric_header_row)
    metric_data_start = metric_header_row + 1
    for m in METRIC_NAMES:
        row = [m]
        for s in snapshots:
            ms = getattr(s.report, m.lower())
            row.append(round(ms.mean, 4))
        ws.append(row)
    metric_data_end = ws.max_row

    # 嵌入 grouped BarChart：6 metrics × N labels
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = f"评估指标 mean 对比 ({n} runs)"
    chart.y_axis.title = "mean"
    chart.x_axis.title = "metric"
    chart.height = 10
    chart.width = 20
    data_ref = Reference(
        ws, min_col=2, max_col=1 + n,
        min_row=metric_header_row, max_row=metric_data_end,
    )
    cat_ref = Reference(
        ws, min_col=1, max_col=1,
        min_row=metric_data_start, max_row=metric_data_end,
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    anchor_col = get_column_letter(n + 3)
    ws.add_chart(chart, f"{anchor_col}{metric_header_row}")

    _autosize(ws)

    # ─── Sheet 2: Metrics (full stats) ───
    ws2 = wb.create_sheet("Metrics")
    ws2.append(["metric", "label", "n", "mean", "median", "std", "min", "max"])
    _style_header(ws2, 1)
    for m in METRIC_NAMES:
        for s in snapshots:
            ms = getattr(s.report, m.lower())
            ws2.append(
                [m, s.label, ms.n,
                 round(ms.mean, 4), round(ms.median, 4), round(ms.std, 4),
                 round(ms.min, 4), round(ms.max, 4)]
            )
    _autosize(ws2)

    # ─── Sheet 3: Runtime ───
    ws3 = wb.create_sheet("Runtime")
    ws3.append([
        "metric", "label", "count", "mean", "p50", "p95", "p99", "min", "max", "sum",
    ])
    _style_header(ws3, 1)
    for s in snapshots:
        r = s.report.e2e_latency_ms
        if r:
            ws3.append([
                r.metric, s.label, r.count,
                round(r.mean, 2), round(r.p50, 2), round(r.p95, 2),
                round(r.p99, 2), round(r.min, 2), round(r.max, 2), round(r.sum, 2),
            ])
    for s in snapshots:
        r = s.report.tokens_per_task
        if r:
            ws3.append([
                r.metric, s.label, r.count,
                round(r.mean, 2), round(r.p50, 2), round(r.p95, 2),
                round(r.p99, 2), round(r.min, 2), round(r.max, 2), round(r.sum, 2),
            ])
    _autosize(ws3)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Renderer — PNG (option B)
# ---------------------------------------------------------------------------


def render_png(
    snapshots: list[RunSnapshot],
    charts_dir: Path,
) -> dict[str, Path]:
    """生成 4 张 PNG 独立图表。matplotlib 缺失时记 warning + 返回空字典。

    md 主报告不引用这些图（用户要求纯文本），但 PNG 仍单独输出，供需要图形可视化
    的用户查看 charts/ 目录直接打开。
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
        import numpy as np
    except ImportError as e:
        logger.warning(f"matplotlib 不可用，跳过 PNG: {e}")
        return {}

    # CJK 字体探测
    cjk_candidates = [
        "PingFang SC", "Heiti SC", "STHeiti", "Hiragino Sans GB",
        "Arial Unicode MS", "Microsoft YaHei", "SimHei",
        "Noto Sans CJK SC", "Source Han Sans SC", "WenQuanYi Zen Hei",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    chosen = next((c for c in cjk_candidates if c in available), None)
    if chosen:
        matplotlib.rcParams["font.sans-serif"] = [chosen] + list(
            matplotlib.rcParams.get("font.sans-serif", [])
        )
        matplotlib.rcParams["axes.unicode_minus"] = False
    else:
        logger.warning("未找到 CJK 字体，PNG 中文将显示为 □")

    charts_dir.mkdir(parents=True, exist_ok=True)
    chart_paths: dict[str, Path] = {}

    n = len(snapshots)
    labels = [s.label for s in snapshots]

    # 1) Grouped bar: 0/1 指标对比（Eff / CEI 不在 [0,1]，单独画避免轴失真）
    bounded_names = ("ECR", "TS", "IFS", "IISR", "AR")
    indices = np.arange(len(bounded_names))
    bar_width = 0.8 / max(n, 1)

    fig, ax = plt.subplots(figsize=(max(10, len(bounded_names) * 1.5), 6))
    for i, s in enumerate(snapshots):
        means = [getattr(s.report, m.lower()).mean for m in bounded_names]
        stds = [getattr(s.report, m.lower()).std for m in bounded_names]
        offset = (i - (n - 1) / 2) * bar_width
        ax.bar(
            indices + offset, means, bar_width,
            yerr=stds, capsize=3, label=s.label, alpha=0.85,
        )
    ax.set_xticks(indices)
    ax.set_xticklabels(bounded_names)
    ax.set_ylim(0, 1.1)
    ax.set_ylabel("mean")
    ax.set_title(f"0/1 指标 mean 对比 ({n} runs)")
    ax.legend(loc="upper right", fontsize=8)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    p = charts_dir / "metrics_grouped_bar.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    chart_paths["metrics_grouped_bar"] = p

    # 2) Boxplot: 0/1 指标 6 子图，每图 N 个 box
    fig, axes = plt.subplots(2, 3, figsize=(15, 9))
    for ax, m in zip(axes.flat, bounded_names):
        per_run_data: list[list[float]] = []
        labels_filtered: list[str] = []
        for s in snapshots:
            data = [float(r.metrics.get(m, 0.0) or 0.0) for r in s.rows]
            if data:
                per_run_data.append(data)
                labels_filtered.append(s.label)
        if per_run_data:
            ax.boxplot(per_run_data, labels=labels_filtered, showmeans=True)
        ax.set_title(m)
        ax.set_ylim(-0.05, 1.05)
        ax.tick_params(axis="x", rotation=15, labelsize=8)
        ax.grid(axis="y", alpha=0.3)
    fig.suptitle(f"per-case 分布对比 ({n} runs)", fontsize=14)
    fig.tight_layout()
    p = charts_dir / "metrics_grouped_box.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    chart_paths["metrics_grouped_box"] = p

    # 3) Runtime compare：TTFT mean/p95 + Tokens mean
    if any(s.report.e2e_latency_ms or s.report.tokens_per_task for s in snapshots):
        fig, axes = plt.subplots(1, 2, figsize=(14, 5))
        x = np.arange(n)

        ttft_means = [s.report.e2e_latency_ms.mean if s.report.e2e_latency_ms else 0 for s in snapshots]
        ttft_p95 = [s.report.e2e_latency_ms.p95 if s.report.e2e_latency_ms else 0 for s in snapshots]
        axes[0].bar(x - 0.2, ttft_means, 0.4, label="mean", color="#4472C4")
        axes[0].bar(x + 0.2, ttft_p95, 0.4, label="p95", color="#ED7D31")
        axes[0].set_xticks(x)
        axes[0].set_xticklabels(labels, rotation=15, fontsize=8)
        axes[0].set_ylabel("ms")
        axes[0].set_title("E2E_Latency mean / p95")
        axes[0].legend()
        axes[0].grid(axis="y", alpha=0.3)

        tok_means = [
            s.report.tokens_per_task.mean if s.report.tokens_per_task else 0
            for s in snapshots
        ]
        axes[1].bar(x, tok_means, 0.6, color="#5B9BD5", alpha=0.85)
        axes[1].set_xticks(x)
        axes[1].set_xticklabels(labels, rotation=15, fontsize=8)
        axes[1].set_ylabel("tokens / task")
        axes[1].set_title("Tokens mean per task")
        if tok_means and max(tok_means) > 0:
            for i, v in enumerate(tok_means):
                axes[1].text(
                    i, v + max(tok_means) * 0.02,
                    f"{v:.0f}", ha="center", fontsize=8,
                )
        axes[1].grid(axis="y", alpha=0.3)

        fig.tight_layout()
        p = charts_dir / "runtime_compare.png"
        fig.savefig(p, dpi=140)
        plt.close(fig)
        chart_paths["runtime_compare"] = p

    # 4) Status compare：堆叠条
    fig, ax = plt.subplots(figsize=(max(8, n * 1.5), 5))
    x = np.arange(n)
    bottom = np.zeros(n)
    statuses = [
        ("success", [s.counts.n_success for s in snapshots], "#70AD47"),
        ("skipped", [s.counts.n_skipped for s in snapshots], "#FFC000"),
        ("error", [s.counts.n_error for s in snapshots], "#C00000"),
        ("unparseable", [s.counts.n_unparseable for s in snapshots], "#7F7F7F"),
    ]
    for name, vals, color in statuses:
        ax.bar(x, vals, bottom=bottom, label=name, color=color, alpha=0.85)
        bottom = bottom + np.array(vals)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=15, fontsize=9)
    ax.set_ylabel("count")
    ax.set_title("status counts per run")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    p = charts_dir / "status_compare.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    chart_paths["status_compare"] = p

    return chart_paths


# ===========================================================================
# Grouped (per-model) comparison —— 仅均值，双层报告
# ===========================================================================
#
# 与上面的 compare() 互补：那个把每份 csv 当独立 run、内部用 metrics_summary
# 的完整 stats（mean/median/std/...）；这个把多份 csv 按 model 分组，**只算
# 均值**，输出双层 markdown：
#
#   ① 单模型多轮均值（顶部）—— 每个 model 一行，把该 model 所有 runs 的
#      所有 success cases 拼成一个大集合再算均值
#   ② 单模型每轮明细（分组明细）—— 每 model 一节，节内每 run 一行
#
# 不依赖 metrics_summary.aggregate_batch；走 compute_run_means（mean-only），
# 完全在 service 层自建。


@dataclass
class _ModelRunEntry:
    """单 model 下的某一份 csv：行级数据 + 计数 + per-run 均值。"""
    csv_path: Path
    counts: StatusCounts
    means: RunMeans
    n_rows_loaded: int  # = len(rows)，便于审计


@dataclass
class _ModelGroup:
    """一个模型的整组聚合视图：包含每 run 明细 + 跨 runs 合并均值。"""
    model: str
    runs: list[_ModelRunEntry]
    merged: RunMeans  # 跨 runs 跨 cases 的算术平均
    total_counts: StatusCounts  # n_total / n_success / ... 之和


def _aggregate_status_counts(items: Iterable[StatusCounts]) -> StatusCounts:
    out = StatusCounts()
    for c in items:
        out.n_total += c.n_total
        out.n_success += c.n_success
        out.n_skipped += c.n_skipped
        out.n_error += c.n_error
        out.n_unparseable += c.n_unparseable
    return out


def _build_model_group(model: str, paths: list[Path]) -> _ModelGroup:
    """对单模型的多个 csv 跑 load + per-run means + merged means。"""
    runs: list[_ModelRunEntry] = []
    all_rows: list[EvalRow] = []
    for p in paths:
        rows, counts = load_eval_results(p)
        all_rows.extend(rows)
        runs.append(_ModelRunEntry(
            csv_path=p,
            counts=counts,
            means=compute_run_means(rows, n_runs=1),
            n_rows_loaded=len(rows),
        ))
    merged = compute_run_means(all_rows, n_runs=len(paths))
    total_counts = _aggregate_status_counts(r.counts for r in runs)
    return _ModelGroup(model=model, runs=runs, merged=merged, total_counts=total_counts)


def _format_metric(v: float | int | None) -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        # 6 个 [0,1] 指标用 4 位小数；ttft/tokens 这种由调用端格式化
        return f"{v:.4f}"
    return str(v)


def _md_table(headers: list[str], rows: list[list[str]]) -> str:
    out = ["| " + " | ".join(headers) + " |",
           "| " + " | ".join(["---"] * len(headers)) + " |"]
    for r in rows:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def _means_row(label: str, n_runs_str: str, m: RunMeans) -> list[str]:
    """RunMeans → 表格一行字符串数组。"""
    return [
        label,
        n_runs_str,
        str(m.n_cases),
        _format_metric(m.ecr),
        _format_metric(m.ts),
        _format_metric(m.ifs),
        _format_metric(m.iisr),
        _format_metric(m.ar),
        (f"{m.eff:.4f} (n={m.n_eff})" if m.eff is not None else "—"),
        (f"{m.ses:.4f}" if m.ses is not None else "—"),
        (f"{m.cei:.4f}" if m.cei is not None else "—"),
        f"{m.e2e_latency_ms:.2f}",
        f"{m.input_tokens:.1f}",
        f"{m.output_tokens:.1f}",
        f"{m.total_tokens:.1f}",
    ]


_MEANS_HEADERS = [
    "Label", "n_runs", "n_cases", "ECR", "TS", "IFS", "IISR", "AR", "Eff", "SES", "CEI",
    "E2E_Latency(ms)", "InputTokens", "OutputTokens", "TotalTokens",
]


def render_grouped_md(groups: list[_ModelGroup], md_path: Path, run_ts: str) -> None:
    """渲染双层 markdown：① 单模型多轮均值；② 单模型每轮明细。"""
    lines: list[str] = []
    lines.append(f"# Eval Comparison (means only) — {run_ts}")
    lines.append("")
    lines.append(
        "E2E_Latency.mean = mean of per-case `avg_ttft_ms`（mean-of-means）；"
        "每个 case 的 `avg_ttft_ms` 已是 turn-level 均值（sum/n_assistant），"
        "展示层统一标签为 E2E_Latency。"
    )
    lines.append("")

    # ── ① 顶部：单模型多轮均值 ────────────────────────────────────────
    lines.append("## ① 单模型多轮均值（汇总）")
    lines.append("")
    rows = []
    for g in groups:
        rows.append(_means_row(g.model, str(len(g.runs)), g.merged))
    lines.append(_md_table(_MEANS_HEADERS, rows))
    lines.append("")

    # 状态计数小表
    lines.append("### 样本状态统计（按 model 合并）")
    lines.append("")
    status_headers = ["Model", "total", "success", "skipped", "error", "unparseable"]
    status_rows = []
    for g in groups:
        c = g.total_counts
        status_rows.append([
            g.model, str(c.n_total), str(c.n_success), str(c.n_skipped),
            str(c.n_error), str(c.n_unparseable),
        ])
    lines.append(_md_table(status_headers, status_rows))
    lines.append("")

    # ── ② 每模型每轮明细 ─────────────────────────────────────────────
    lines.append("## ② 单模型每轮评测明细")
    lines.append("")
    for g in groups:
        lines.append(f"### {g.model}（{len(g.runs)} runs）")
        lines.append("")
        run_rows = []
        for entry in g.runs:
            # 每 run 行：用 csv 文件名（不含目录、不含扩展名）当 label
            label = entry.csv_path.stem
            run_rows.append(_means_row(label, "1", entry.means))
        lines.append(_md_table(_MEANS_HEADERS, run_rows))
        lines.append("")

        # per-run 状态小表
        per_run_status_headers = ["Run", "total", "success", "skipped", "error", "unparseable"]
        per_run_status_rows = []
        for entry in g.runs:
            c = entry.counts
            per_run_status_rows.append([
                entry.csv_path.stem, str(c.n_total), str(c.n_success),
                str(c.n_skipped), str(c.n_error), str(c.n_unparseable),
            ])
        lines.append(_md_table(per_run_status_headers, per_run_status_rows))
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")


def render_grouped_json(
    groups: list[_ModelGroup], json_path: Path, run_ts: str
) -> None:
    """渲染机器消费用的 JSON，对称 markdown 结构。"""
    payload = {
        "run_ts": run_ts,
        "n_models": len(groups),
        "models": [
            {
                "model": g.model,
                "n_runs": len(g.runs),
                "merged": g.merged.as_dict(),
                "total_counts": {
                    "n_total": g.total_counts.n_total,
                    "n_success": g.total_counts.n_success,
                    "n_skipped": g.total_counts.n_skipped,
                    "n_error": g.total_counts.n_error,
                    "n_unparseable": g.total_counts.n_unparseable,
                },
                "runs": [
                    {
                        "csv": str(entry.csv_path),
                        "label": entry.csv_path.stem,
                        "means": entry.means.as_dict(),
                        "counts": {
                            "n_total": entry.counts.n_total,
                            "n_success": entry.counts.n_success,
                            "n_skipped": entry.counts.n_skipped,
                            "n_error": entry.counts.n_error,
                            "n_unparseable": entry.counts.n_unparseable,
                        },
                    }
                    for entry in g.runs
                ],
            }
            for g in groups
        ],
    }
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def compare_grouped(
    groups: dict[str, list[Path]],
    output_dir: Path,
) -> ComparisonProducts:
    """按 model 分组、仅均值的对比报告。

    与 ``compare`` 的区别：
      - 输入是 model → list[csv path] 映射；同 model 多个 csv 视为多次评测
      - 输出**只有** markdown + json，无 xlsx 内嵌图表、无 matplotlib png
      - 指标只算算术平均值，不算 median/std/percentile

    Args:
        groups: ``{model_name: [csv_path, ...]}``，至少 1 个 model
        output_dir: 报告根目录；产物落 ``<output_dir>/compare/<ts>/``

    Returns:
        ``ComparisonProducts``（``xlsx_path`` / ``charts_dir`` 字段为 None）
    """
    if not groups:
        raise ValueError("compare_grouped 至少需要 1 个 model")

    model_groups = [_build_model_group(m, paths) for m, paths in groups.items()]

    for g in model_groups:
        logger.info(
            "[%s] runs=%d, total_cases=%d, success=%d (skipped=%d, error=%d, unparseable=%d)",
            g.model, len(g.runs), g.total_counts.n_total, g.total_counts.n_success,
            g.total_counts.n_skipped, g.total_counts.n_error, g.total_counts.n_unparseable,
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = output_dir / "compare" / ts
    base_dir.mkdir(parents=True, exist_ok=True)

    md_path = base_dir / "report.md"
    json_path = base_dir / "comparison_summary.json"
    render_grouped_md(model_groups, md_path, ts)
    render_grouped_json(model_groups, json_path, ts)

    # 控制台输出 md，便于跑完直接看
    print(md_path.read_text(encoding="utf-8"))

    return ComparisonProducts(
        out_dir=base_dir,
        json_path=json_path,
        md_path=md_path,
        xlsx_path=None,
        charts_dir=None,
        chart_paths={},
    )
