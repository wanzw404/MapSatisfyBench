import logging
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.schemas.batch_evaluate import EvaluateCase, EvaluateResult
from app.schemas.dialogue_simulator import DialogueCase
from app.paths import INPUT_DIR, OUTPUT_DIR

logger = logging.getLogger(__name__)

INPUT_DIR = Path(__file__).parent.parent.parent / "data" / "inputs"
OUTPUT_DIR = Path(__file__).parent.parent.parent / "data" / "outputs" / "simulator_res"

# 输入列名（支持中英文别名映射）
INPUT_COLS = ["query", "location", "tool", "expected"]
REQUIRED_COLS = ["query"]

# 批量评测的列定义。case_index 在最前用于事后排序回原始顺序（并发完成时
# xlsx 行序与输入顺序不一致）。
INDEX_COL = "case_index"
BATCH_RESULT_HEADERS = (
    [INDEX_COL]
    + ["query", "location", "tool", "expected"]
    + ["conversation_id", "status", "final_response", "tool_calls", "execution_time_ms", "error_message"]
)

# 列宽：A=case_index, B=query, ..., K=error_message
BATCH_RESULT_COLUMN_WIDTHS = {
    "A": 10, "B": 60, "C": 22, "D": 20, "E": 60,
    "F": 36, "G": 10, "H": 80, "I": 80, "J": 14, "K": 50,
}

# 输出追加列名（保留旧别名以便其它代码引用）
OUTPUT_COLS = [
    "conversation_id",
    "status",
    "final_response",
    "tool_calls",
    "execution_time_ms",
    "error_message",
]


def read_cases(filename: str) -> list[EvaluateCase]:
    """读取 /data/inputs 下的 Excel 文件，返回评测用例列表。"""
    file_path = INPUT_DIR / filename
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = load_workbook(file_path)
    ws = wb.active

    # 读取表头
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    header_map = {h.lower(): idx for idx, h in enumerate(headers)}

    # 校验必填列
    missing = [c for c in REQUIRED_COLS if c not in header_map]
    if missing:
        raise ValueError(f"Excel 缺少必填列: {missing}")

    cases: list[EvaluateCase] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        query = _get_cell(row, header_map, "query")
        if not query:
            continue

        cases.append(
            EvaluateCase(
                query=query,
                location=_get_cell(row, header_map, "location"),
                tool=_get_cell(row, header_map, "tool"),
                expected=_get_cell(row, header_map, "expected"),
            )
        )

    logger.info(f"从 {filename} 读取到 {len(cases)} 条用例")
    return cases


def write_results(filename: str, results: list[EvaluateResult]) -> Path:
    """将执行结果写入 /data/outputs/simulator/{filename}_result_{timestamp}.xlsx。"""
    from datetime import datetime

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(filename).stem
    output_name = f"{stem}_result_{timestamp}.xlsx"
    output_path = OUTPUT_DIR / output_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # 表头：输入列 + 输出列
    all_headers = INPUT_COLS + OUTPUT_COLS
    ws.append(all_headers)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 写入数据行
    for r in results:
        ws.append([
            r.query,
            r.location or "",
            r.tool or "",
            r.expected or "",
            r.conversation_id,
            r.status,
            r.final_response,
            str(r.tool_calls),
            r.execution_time_ms,
            r.error_message or "",
        ])

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)
    logger.info(f"结果已写入: {output_path}")
    return output_path


def read_dialogue_cases(filename: str) -> list[DialogueCase]:
    """读取 /data/inputs 下的 Excel 文件，返回多轮对话评测用例列表。"""
    file_path = INPUT_DIR / filename
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = load_workbook(file_path)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    header_map = {h.lower(): idx for idx, h in enumerate(headers)}

    missing = [c for c in REQUIRED_COLS if c not in header_map]
    if missing:
        raise ValueError(f"Excel 缺少必填列: {missing}")

    cases: list[DialogueCase] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        query = _get_cell(row, header_map, "query")
        if not query:
            continue

        cases.append(
            DialogueCase(
                task_id=_get_cell(row, header_map, "task_id"),
                query=query,
                context=_get_cell(row, header_map, "context"),
                time=_get_cell(row, header_map, "time"),
                location=_get_cell(row, header_map, "location"),
                tool=_get_cell(row, header_map, "tool"),
                persona=_get_cell(row, header_map, "persona"),
                full_intent=_get_cell(row, header_map, "full_intent"),
                expected=_get_cell(row, header_map, "expected"),
                ground_truth=_get_cell(row, header_map, "ground_truth"),
            )
        )

    logger.info(f"从 {filename} 读取到 {len(cases)} 条对话用例")
    return cases


def _get_cell(row: tuple, header_map: dict, key: str) -> str | None:
    """从行数据中提取指定列的值。"""
    idx = header_map.get(key)
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip()


def create_batch_output_file(input_filename: str) -> Path:
    """为批量评测预先创建带表头的空 xlsx，返回路径。

    BatchResultWriter 之后会 load 这个文件做流式 append，避免每条 case 都
    重新打开/重写整文件。
    """
    from datetime import datetime

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(input_filename).stem
    output_name = f"{stem}_result_{timestamp}.xlsx"
    output_path = OUTPUT_DIR / output_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(BATCH_RESULT_HEADERS)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 固定列宽（流式写入时不能再 auto-fit）
    for col_letter, width in BATCH_RESULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)
    logger.info(f"已创建批量评测结果文件: {output_path}")
    return output_path
