"""Eval result summary service.

把 ``data/outputs/evaluation_res/*.xlsx``（来自
``batch_evaluate_from_simulator``）做一次批次级统计 + 落盘 + 可视化。

设计原则：
- 完全复用 ``app.core.evaluation.metrics_summary`` 的 ``aggregate_batch`` /
  ``format_batch_report``，**不修改其逻辑**。本模块只做：xlsx → MetricScores
  列表 → 调底层聚合 → 渲染。
- 所有业务逻辑在本模块；CLI 层（``app/scripts/aggregate_eval_results.py``）
  仅是 argparse + 调用入口，便于将来 HTTP 路由直接 import 复用。
"""

from __future__ import annotations

import csv
import json
import logging
import sys

# csv 默认单字段上限 128KB；评测 CSV 的 results 列是 JSON-dumped MetricScores
# + details，含失败诊断细节时可能长。提到 sys.maxsize 避免 _csv.Error 撞上限。
csv.field_size_limit(sys.maxsize)
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Literal, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.evaluation.metrics_summary import (
    BatchReport,
    aggregate_batch,
    format_batch_report,
)
from app.core.evaluation.schema import MetricScores

logger = logging.getLogger(__name__)


METRIC_NAMES: tuple[str, ...] = ("ECR", "TS", "IFS", "IISR", "AR", "Eff", "SES", "CEI")
VisualizeMode = Literal["xlsx", "png", "both", "none"]


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------


@dataclass
class EvalRow:
    """一行成功 case 的扁平视图（对应 xlsx 一行 + parse 后的 metrics）。"""

    case_id: str
    status: str
    metrics: dict[str, float]    # ECR/TS/IFS/IISR/AR/Eff/CEI + avg_ttft_ms + tokens
    raw_results: dict[str, Any]  # 原始 results JSON（保留 details 等）

    @property
    def e2e_latency_ms(self) -> float:
        """End-to-end latency (ms) per case；新 CSV 写 ``e2e_latency_ms``，
        历史 CSV 字段名是 ``avg_ttft_ms``，这里做透明兼容读取。"""
        v = self.metrics.get("e2e_latency_ms")
        if v is None:
            v = self.metrics.get("avg_ttft_ms", 0.0)
        return float(v or 0.0)

    @property
    def input_tokens(self) -> int:
        """Session-level prompt tokens；老 CSV 缺该字段返 0（不再 fallback 到
        total_tokens 一半之类的猜测——分开统计本来就是为了精确）。"""
        return int(self.metrics.get("input_tokens", 0) or 0)

    @property
    def output_tokens(self) -> int:
        """Session-level completion tokens；老 CSV 缺该字段返 0。"""
        return int(self.metrics.get("output_tokens", 0) or 0)

    @property
    def total_tokens(self) -> int:
        """优先取 results.metrics.total_tokens；如果该字段也缺（极旧），
        fallback 用 input + output 兜底。"""
        v = self.metrics.get("total_tokens")
        if v is not None:
            return int(v or 0)
        return self.input_tokens + self.output_tokens


@dataclass
class StatusCounts:
    """xlsx 中各 status 的 case 数。"""

    n_total: int = 0
    n_success: int = 0
    n_skipped: int = 0
    n_error: int = 0
    n_unparseable: int = 0   # results 列存在但 JSON 解析失败


@dataclass
class SummaryProducts:
    """`summarize` 的产物：所有落盘文件的路径。"""

    json_path: Path
    txt_path: Path
    xlsx_path: Optional[Path] = None
    charts_dir: Optional[Path] = None
    chart_paths: dict[str, Path] = field(default_factory=dict)
    html_path: Optional[Path] = None


@dataclass
class RunMeans:
    """单 run / 多 run 合并视角下的「仅均值」聚合结果。

    报告精简后用——只保留 6 指标 + TTFT + Tokens 的算术平均，舍去
    median / std / percentile 这些次级统计量。

    构造规则（2026-05-26 分母策略）：
      - 6 指标的 mean = **全部 case 的算术平均**（含失败 case，按 0 计入）
        ``n_cases`` = 所有传入行（成功 + skipped + error + unparseable）
      - Eff 跳过 ``details.Eff.skipped`` 与 ``Eff is None`` 的 case（数学
        不可计算），失败 case 因 zero_eval_row 设了 Eff=0.0 仍计入
      - **TTFT.mean / Tokens.mean = 仅 success case 的算术平均**
        runtime 物理量缺测不补 0，避免污染 latency 均值
      - n_runs  = 合并了几份 CSV（单 run = 1）
    """

    n_runs: int = 0
    n_cases: int = 0
    ecr: float = 0.0
    ts: float = 0.0
    ifs: float = 0.0
    iisr: float = 0.0
    ar: float = 0.0
    eff: Optional[float] = None  # 全部 case 都没 Eff 时为 None
    ses: Optional[float] = None  # 全部 case 都没 SES 时为 None
    cei: Optional[float] = None  # 全部 case 都没 CEI 时为 None
    n_eff: int = 0  # 参与 Eff 平均的 case 数（≤ n_cases）
    e2e_latency_ms: float = 0.0  # 案上数据来源 metrics.e2e_latency_ms（老 CSV 兼容读 avg_ttft_ms）
    input_tokens: float = 0.0   # session-level prompt tokens 的 case-mean
    output_tokens: float = 0.0  # session-level completion tokens 的 case-mean
    total_tokens: float = 0.0   # 保留作向后兼容；= input + output（来自原始 metrics 字段）

    def as_dict(self) -> dict[str, Any]:
        return {
            "n_runs": self.n_runs,
            "n_cases": self.n_cases,
            "ECR": round(self.ecr, 4),
            "TS": round(self.ts, 4),
            "IFS": round(self.ifs, 4),
            "IISR": round(self.iisr, 4),
            "AR": round(self.ar, 4),
            "Eff": (round(self.eff, 4) if self.eff is not None else None),
            "SES": (round(self.ses, 4) if self.ses is not None else None),
            "CEI": (round(self.cei, 4) if self.cei is not None else None),
            "n_Eff_cases": self.n_eff,
            "E2E_Latency_mean_ms": round(self.e2e_latency_ms, 2),
            "InputTokens_mean": round(self.input_tokens, 1),
            "OutputTokens_mean": round(self.output_tokens, 1),
            "Tokens_mean": round(self.total_tokens, 1),
        }


def compute_run_means(
    rows: Iterable[EvalRow],
    *,
    n_runs: int = 1,
) -> RunMeans:
    """对一组 EvalRow 计算「仅均值」聚合。

    输入是 ``load_eval_results`` 返回的全部行（含 status=success / skipped /
    error / unparseable）。6 指标的分母 = 全部行数；失败 case 按 0 分计入。
    runtime / Eff 走 success-only 路径（见 RunMeans docstring）。

    多 run 合并：把多个 CSV 的 rows 拼成一个大 list 再传进来即可——
    所有指标都是跨所有 case 的算术平均，与「先按 run 平均再按 model 平均」
    会得到不同结果（前者是 case-level mean、后者是 mean-of-means-of-means），
    本函数采用前者，case-level mean 是用户期望的语义。
    """
    rows_list = list(rows)
    n = len(rows_list)
    if n == 0:
        return RunMeans(n_runs=n_runs, n_cases=0)

    # 6 指标：全量分母（失败 case 的 metrics 已被 _zero_eval_row 置 0）
    def _mean(key: str) -> float:
        vals = [float(r.metrics.get(key, 0.0) or 0.0) for r in rows_list]
        return sum(vals) / len(vals) if vals else 0.0

    # Eff: 跳过 None / details.Eff.skipped（数学不可计算）；失败 case 因
    # _zero_eval_row 设了 Eff=0.0，会进入均值。SES 同理。
    eff_vals: list[float] = []
    ses_vals: list[float] = []
    cei_vals: list[float] = []
    for r in rows_list:
        v = r.metrics.get("Eff")
        if v is None:
            continue
        details_eff = (r.raw_results or {}).get("details", {}).get("Eff", {})
        if isinstance(details_eff, dict) and details_eff.get("skipped"):
            continue
        eff_vals.append(float(v))
        # SES follows Eff skip logic
        ses_v = r.metrics.get("SES")
        if ses_v is not None:
            ses_vals.append(float(ses_v))
        # CEI follows SES skip logic
        cei_v = r.metrics.get("CEI")
        if cei_v is not None:
            cei_vals.append(float(cei_v))

    # runtime（E2E_Latency / tokens）：仅 success case 参与；缺测不补 0 以免污染均值
    success_rows = [r for r in rows_list if r.status == "success"]
    ttft_vals = [r.e2e_latency_ms for r in success_rows]
    input_vals = [r.input_tokens for r in success_rows]
    output_vals = [r.output_tokens for r in success_rows]
    total_vals = [r.total_tokens for r in success_rows]

    return RunMeans(
        n_runs=n_runs,
        n_cases=n,
        ecr=_mean("ECR"),
        ts=_mean("TS"),
        ifs=_mean("IFS"),
        iisr=_mean("IISR"),
        ar=_mean("AR"),
        eff=(sum(eff_vals) / len(eff_vals)) if eff_vals else None,
        ses=(sum(ses_vals) / len(ses_vals)) if ses_vals else None,
        cei=(sum(cei_vals) / len(cei_vals)) if cei_vals else None,
        n_eff=len(eff_vals),
        e2e_latency_ms=sum(ttft_vals) / len(ttft_vals) if ttft_vals else 0.0,
        input_tokens=sum(input_vals) / len(input_vals) if input_vals else 0.0,
        output_tokens=sum(output_vals) / len(output_vals) if output_vals else 0.0,
        total_tokens=sum(total_vals) / len(total_vals) if total_vals else 0.0,
    )


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def _iter_xlsx_rows(xlsx_path: Path) -> Iterator[dict[str, Any]]:
    """读 xlsx → 逐行 dict。

    NOTE: 不能用 read_only=True。BatchResultWriter / batch_evaluate_from_simulator
    流式写入产出的 xlsx，在 read-only 模式下 openpyxl 经常只看到表头行，所有数据行
    被跳过 → 统计全 0。改用普通模式（评测结果通常 < 几 MB，全量加载 OK）。
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    if "case_id" not in headers:
        wb.close()
        raise ValueError(f"{xlsx_path} 缺少 case_id 列")
    for row in rows_iter:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        yield {h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)}
    wb.close()


def _iter_csv_rows(csv_path: Path) -> Iterator[dict[str, Any]]:
    """读 csv → 逐行 dict（utf-8-sig 自动剥 BOM）。"""
    with open(csv_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None or "case_id" not in reader.fieldnames:
            raise ValueError(f"{csv_path} 缺少 case_id 列")
        for row in reader:
            if not row or all(
                v is None or str(v).strip() == "" for v in row.values()
            ):
                continue
            yield row


def _zero_eval_row(case_id: str, status: str) -> EvalRow:
    """失败 / skipped / unparseable 行的零分占位。

    ECR/TS/IFS/IISR/AR/Eff/CEI 全 0；e2e_latency_ms / token 字段也置 0
    但下游 ``compute_run_means`` 与 ``aggregate_batch`` runtime stats 会
    按 status != 'success' 过滤掉，不会污染 latency / token 均值。
    """
    return EvalRow(
        case_id=case_id,
        status=status,
        metrics={
            "ECR": 0.0, "TS": 0.0, "IFS": 0.0, "IISR": 0.0, "AR": 0.0,
            # Eff 用 0.0（不是 None）让失败 case 进 Eff 均值。
            # "数学不可计算"（GT max_allowed=0）走 details.Eff.skipped 路径，
            # 与失败是不同概念。SES 同理。
            "Eff": 0.0,
            "SES": 0.0,
            # CEI 与其它 0/1 指标一致：失败 case 按 0 计入分母
            "CEI": 0.0,
            "e2e_latency_ms": 0.0,
            "input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0,
        },
        raw_results={},
    )


def load_eval_results(path: Path) -> tuple[list[EvalRow], StatusCounts]:
    """读评测结果文件并解析每行 results 列。

    支持两种格式（按后缀分流）：
      - ``.csv``  → 新版产出格式（``EvaluationResultWriter`` 写出，并发即写）
      - ``.xlsx`` → 历史评测产物，向后兼容继续可读

    分母策略（2026-05-26）：所有非 success 行（skipped / error / unparseable）
    都构造零分 EvalRow 入列，``EvalRow.status`` 保留原状态以便下游路由：
      * 0/1 指标均值 → 计入分母（按 0 分参与）
      * runtime (TTFT / tokens) 均值 → 仅看 status='success' 行
      * outlier top/bottom → 仅看 status='success' 行
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows_iter = _iter_csv_rows(path)
    elif suffix in (".xlsx", ".xlsm"):
        rows_iter = _iter_xlsx_rows(path)
    else:
        raise ValueError(
            f"不支持的评测结果格式: {path.suffix} (期望 .csv 或 .xlsx)"
        )

    eval_rows: list[EvalRow] = []
    counts = StatusCounts()

    for row in rows_iter:
        counts.n_total += 1

        case_id = str(row.get("case_id") or "").strip()
        status = str(row.get("status") or "").strip().lower()

        if status == "skipped":
            counts.n_skipped += 1
            eval_rows.append(_zero_eval_row(case_id, "skipped"))
            continue
        if status == "error":
            counts.n_error += 1
            eval_rows.append(_zero_eval_row(case_id, "error"))
            continue

        # status == "success" or 空（旧 xlsx）— 尝试解析 results
        results_cell = row.get("results")
        if not results_cell or not str(results_cell).strip():
            counts.n_unparseable += 1
            logger.warning(f"[{case_id}] results 列为空，按 0 分计入均值")
            eval_rows.append(_zero_eval_row(case_id, "unparseable"))
            continue

        try:
            parsed = json.loads(str(results_cell))
        except (json.JSONDecodeError, TypeError) as e:
            counts.n_unparseable += 1
            logger.warning(f"[{case_id}] results JSON 解析失败 ({e}), 按 0 分计入均值")
            eval_rows.append(_zero_eval_row(case_id, "unparseable"))
            continue

        # 标准结构：{"metrics": {ECR:..., TS:..., ..., avg_ttft_ms, total_tokens}, "details": {...}}
        metrics_dict = parsed.get("metrics", {}) if isinstance(parsed, dict) else {}
        if not isinstance(metrics_dict, dict):
            counts.n_unparseable += 1
            logger.warning(f"[{case_id}] results.metrics 不是 dict，按 0 分计入均值")
            eval_rows.append(_zero_eval_row(case_id, "unparseable"))
            continue

        # 兼容旧产物：metrics 里只有 IISR、没有 AR——把 IISR 当作 AR 读入，
        # 同时打 WARNING 提示（数学上 AR=ECR·IISR ≠ IISR，但旧产物已无法
        # 重算 AR，只能直接用 IISR 替代展示）。
        if "AR" not in metrics_dict and "IISR" in metrics_dict:
            metrics_dict = dict(metrics_dict)  # 不污染原 parsed
            metrics_dict["AR"] = metrics_dict["IISR"]
            if not getattr(load_eval_results, "_iisr_warning_shown", False):
                logger.warning(
                    "[%s] 旧版 metrics 仅含 IISR、缺 AR；本 loader 把 IISR 当作 AR "
                    "读入展示。注意：AR=ECR·IISR ≠ IISR，旧产物无法重算 AR；"
                    "本提示只打印一次。", case_id,
                )
                load_eval_results._iisr_warning_shown = True  # type: ignore

        # 必需指标：ECR/TS/IFS 缺任一即按 unparseable 处理；Eff/AR/IISR 可缺
        # （Eff 本就 nullable，极旧产物可能没 AR/IISR 任一）
        # 兼容旧 CSV：接受 "ICR" 作为 "ECR" 的别名
        if "ICR" in metrics_dict and "ECR" not in metrics_dict:
            metrics_dict = dict(metrics_dict)
            metrics_dict["ECR"] = metrics_dict.pop("ICR")
        REQUIRED = ("ECR", "TS", "IFS")
        missing_required = [m for m in REQUIRED if m not in metrics_dict]
        if missing_required:
            counts.n_unparseable += 1
            logger.warning(
                f"[{case_id}] results.metrics 缺必需字段 {missing_required}，按 0 分计入均值"
            )
            eval_rows.append(_zero_eval_row(case_id, "unparseable"))
            continue

        counts.n_success += 1
        # 老 CSV 可能没 IISR / CEI 字段；按 0 兜底，避免下游 None 计算炸；
        # Eff 单独由 `metrics_dict.get("Eff")` 保留 None 语义（下游 _metric_stats
        # 会跳过 None 计算）
        flat_metrics: dict[str, Any] = {}
        for k in METRIC_NAMES:
            v = metrics_dict.get(k)
            if v is None and k not in ("Eff",):
                v = 0.0
            flat_metrics[k] = v
        eval_rows.append(
            EvalRow(
                case_id=case_id,
                status="success",
                metrics=flat_metrics
                | {
                    # 新 CSV 用 e2e_latency_ms，老 CSV 仍是 avg_ttft_ms；
                    # EvalRow.e2e_latency_ms property 会兼容两种读法
                    "e2e_latency_ms": metrics_dict.get(
                        "e2e_latency_ms",
                        metrics_dict.get("avg_ttft_ms", 0.0),
                    ),
                    # 三个 token 字段都从 metrics_dict 透传——新产物里
                    # input/output/total 都有，老产物缺哪个由 EvalRow
                    # 的 property fallback 处理（input/output→0，
                    # total 缺时再回退 input+output）
                    "input_tokens": metrics_dict.get("input_tokens", 0),
                    "output_tokens": metrics_dict.get("output_tokens", 0),
                    # 缺失时显式留 None，让 EvalRow.total_tokens property
                    # 触发 input+output 兜底（极旧产物可能没 total_tokens）
                    "total_tokens": metrics_dict.get("total_tokens"),
                },
                raw_results=parsed,
            )
        )

    return eval_rows, counts


def build_metric_scores(rows: Iterable[EvalRow]) -> list[MetricScores]:
    """把 EvalRow 列表转为 MetricScores（aggregate_batch 入参）。

    走 pydantic model_validate 确保 [0,1] 区间约束生效；不合法的 case 抛 ValueError。
    """
    out: list[MetricScores] = []
    for r in rows:
        ar_val = float(r.metrics.get("AR", r.metrics.get("IISR", 0.0)) or 0.0)
        iisr_val = float(r.metrics.get("IISR", 0.0) or 0.0)
        out.append(
            MetricScores(
                ECR=float(r.metrics["ECR"]),
                TS=float(r.metrics["TS"]),
                IFS=float(r.metrics["IFS"]),
                IISR=iisr_val,
                AR=ar_val,
                Eff=(
                    float(r.metrics["Eff"]) if r.metrics.get("Eff") is not None else None
                ),
                SES=(
                    float(r.metrics["SES"]) if r.metrics.get("SES") is not None else None
                ),
                CEI=(
                    float(r.metrics["CEI"]) if r.metrics.get("CEI") is not None else None
                ),
                e2e_latency_ms=r.e2e_latency_ms,
                input_tokens=r.input_tokens,
                output_tokens=r.output_tokens,
                total_tokens=r.total_tokens,
                details={},
            )
        )
    return out


# ---------------------------------------------------------------------------
# Outliers
# ---------------------------------------------------------------------------


def find_outliers(
    rows: list[EvalRow], top_n: int
) -> dict[str, dict[str, list[tuple[str, float]]]]:
    """每个指标分别列 top_n 高分 / 低分 case_id。

    仅考虑 ``status='success'`` 的行——失败 case 因被 ``_zero_eval_row``
    统一置 0，否则 bottom 列表会被一堆失败 case 顶满，看不出真正"低分"
    的成功 case。

    Returns:
        {metric: {"top": [(case_id, value), ...], "bottom": [(case_id, value), ...]}}
    """
    if top_n <= 0 or not rows:
        return {}
    success_rows = [r for r in rows if r.status == "success"]
    if not success_rows:
        return {}
    out: dict[str, dict[str, list[tuple[str, float]]]] = {}
    for m in METRIC_NAMES:
        pairs = [(r.case_id, float(r.metrics.get(m, 0.0))) for r in success_rows]
        # 高分 desc; 低分 asc；同分按 case_id 字典序保持稳定
        top_sorted = sorted(pairs, key=lambda p: (-p[1], p[0]))[:top_n]
        bot_sorted = sorted(pairs, key=lambda p: (p[1], p[0]))[:top_n]
        out[m] = {"top": top_sorted, "bottom": bot_sorted}
    return out


# ---------------------------------------------------------------------------
# Renderers — JSON / TXT
# ---------------------------------------------------------------------------


def render_json(
    report: BatchReport,
    counts: StatusCounts,
    outliers: dict[str, dict[str, list[tuple[str, float]]]],
    path: Path,
) -> Path:
    payload = {
        "n_total": counts.n_total,
        "n_success": counts.n_success,
        "n_skipped": counts.n_skipped,
        "n_error": counts.n_error,
        "n_unparseable": counts.n_unparseable,
        "denominator_policy": (
            "ECR/TS/IFS/IISR/AR/SES/CEI mean over n_total (failed cases scored 0); "
            "E2E_Latency / Tokens mean over n_success only"
        ),
        "report": report.as_dict(),
        "outliers": {
            m: {
                "top": [{"case_id": cid, "value": v} for cid, v in d["top"]],
                "bottom": [{"case_id": cid, "value": v} for cid, v in d["bottom"]],
            }
            for m, d in outliers.items()
        },
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path


def render_txt(report: BatchReport, counts: StatusCounts, path: Path) -> Path:
    lines = [
        f"# 评测批次统计  (生成时间 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})",
        "",
        f"  样本: total={counts.n_total}  success={counts.n_success}  "
        f"skipped={counts.n_skipped}  error={counts.n_error}  "
        f"unparseable={counts.n_unparseable}",
        "  分母口径: ECR/TS/IFS/IISR/AR/CEI 均值分母 = total（失败 case 按 0 计入）; "
        "E2E_Latency / Tokens 仅 success case 参与",
        "",
        format_batch_report(report),
    ]
    text = "\n".join(lines)
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)
    return path


# ---------------------------------------------------------------------------
# Renderers — XLSX (option A)
# ---------------------------------------------------------------------------


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def render_xlsx(
    report: BatchReport,
    rows: list[EvalRow],
    outliers: dict[str, dict[str, list[tuple[str, float]]]],
    counts: StatusCounts,
    path: Path,
) -> Path:
    """生成 4-sheet 汇总 xlsx：Summary / PerCase / Outliers / Runtime。"""
    wb = Workbook()

    # Sheet 1: Summary（指标表 + bar chart）
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["sample", "n_total", "n_success", "n_skipped", "n_error", "n_unparseable"])
    ws_sum.append([
        "count", counts.n_total, counts.n_success,
        counts.n_skipped, counts.n_error, counts.n_unparseable,
    ])
    _style_header(ws_sum, 1)

    ws_sum.append([])
    ws_sum.append(["metric", "n", "mean", "median", "std", "min", "max"])
    metric_header_row = ws_sum.max_row
    _style_header(ws_sum, metric_header_row)
    metric_data_start = metric_header_row + 1

    for m in (
        report.ecr, report.ts, report.ifs, report.iisr,
        report.ar, report.eff, report.cei,
    ):
        ws_sum.append([m.metric, m.n, m.mean, m.median, m.std, m.min, m.max])
    metric_data_end = ws_sum.max_row

    # Bar chart of mean per metric
    chart = BarChart()
    chart.title = "评估指标均值（含 IISR / CEI）"
    chart.y_axis.title = "mean"
    chart.x_axis.title = "metric"
    chart.style = 11
    data = Reference(
        ws_sum,
        min_col=3, max_col=3,
        min_row=metric_header_row, max_row=metric_data_end,  # 含表头
    )
    cats = Reference(
        ws_sum, min_col=1, max_col=1,
        min_row=metric_data_start, max_row=metric_data_end,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws_sum.add_chart(chart, f"I{metric_header_row}")

    # Runtime 区
    if report.e2e_latency_ms or report.tokens_per_task:
        ws_sum.append([])
        ws_sum.append([
            "runtime_metric", "count", "mean", "p50", "p95", "p99", "min", "max", "sum",
        ])
        rt_hdr = ws_sum.max_row
        _style_header(ws_sum, rt_hdr)
        if report.e2e_latency_ms:
            r = report.e2e_latency_ms
            ws_sum.append([
                r.metric, r.count, r.mean, r.p50, r.p95, r.p99, r.min, r.max, r.sum,
            ])
        if report.tokens_per_task:
            r = report.tokens_per_task
            ws_sum.append([
                r.metric, r.count, r.mean, r.p50, r.p95, r.p99, r.min, r.max, r.sum,
            ])

    _autosize(ws_sum)

    # Sheet 2: PerCase（status 列让查阅者一眼分清成功 / 失败行）
    ws_pc = wb.create_sheet("PerCase")
    ws_pc.append(
        ["case_id", "status"]
        + list(METRIC_NAMES)
        + ["E2E_Latency_ms", "total_tokens"]
    )
    _style_header(ws_pc, 1)
    for r in rows:
        ws_pc.append(
            [r.case_id, r.status]
            + [round(float(r.metrics.get(m, 0.0) or 0.0), 4) for m in METRIC_NAMES]
            + [round(r.e2e_latency_ms, 2), r.total_tokens]
        )
    _autosize(ws_pc, min_width=12, max_width=40)

    # Sheet 3: Outliers
    ws_ol = wb.create_sheet("Outliers")
    ws_ol.append(["metric", "rank", "side", "case_id", "value"])
    _style_header(ws_ol, 1)
    for m in METRIC_NAMES:
        bucket = outliers.get(m, {})
        for rank, (cid, v) in enumerate(bucket.get("top", []), 1):
            ws_ol.append([m, rank, "top", cid, round(v, 4)])
        for rank, (cid, v) in enumerate(bucket.get("bottom", []), 1):
            ws_ol.append([m, rank, "bottom", cid, round(v, 4)])
    _autosize(ws_ol, min_width=10, max_width=40)

    # Sheet 4: Runtime histogram (bar chart 模拟) — 仅 success 行参与，
    # 避免失败 case 的 0 拉低 TTFT / token 视觉对比
    success_rows = [r for r in rows if r.status == "success"]
    if success_rows and (report.e2e_latency_ms or report.tokens_per_task):
        ws_rt = wb.create_sheet("Runtime")
        ws_rt.append(["case_id", "E2E_Latency_ms", "total_tokens"])
        _style_header(ws_rt, 1)
        for r in success_rows:
            ws_rt.append([r.case_id, round(r.e2e_latency_ms, 2), r.total_tokens])

        # 对 E2E_Latency 和 tokens 各画一个 bar chart
        last_row = ws_rt.max_row
        if report.e2e_latency_ms:
            ch = BarChart()
            ch.title = "Per-case E2E_Latency (ms)"
            ch.style = 11
            ch.height = 8
            ch.width = 18
            ch.add_data(
                Reference(ws_rt, min_col=2, max_col=2, min_row=1, max_row=last_row),
                titles_from_data=True,
            )
            ch.set_categories(
                Reference(ws_rt, min_col=1, max_col=1, min_row=2, max_row=last_row)
            )
            ws_rt.add_chart(ch, "E1")
        if report.tokens_per_task:
            ch = BarChart()
            ch.title = "Per-case Tokens"
            ch.style = 12
            ch.height = 8
            ch.width = 18
            ch.add_data(
                Reference(ws_rt, min_col=3, max_col=3, min_row=1, max_row=last_row),
                titles_from_data=True,
            )
            ch.set_categories(
                Reference(ws_rt, min_col=1, max_col=1, min_row=2, max_row=last_row)
            )
            ws_rt.add_chart(ch, "E18")
        _autosize(ws_rt, min_width=12, max_width=40)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Renderers — PNG (option B)
# ---------------------------------------------------------------------------


def render_png(
    report: BatchReport,
    rows: list[EvalRow],
    counts: StatusCounts,
    output_dir: Path,
    *,
    html_dir: Optional[Path] = None,
) -> tuple[dict[str, Path], Optional[Path]]:
    """生成 4 张 PNG + 1 个 HTML 一页式报告。matplotlib 缺失时记 warning 并返回空。

    Args:
        output_dir: PNG 落盘目录（典型 ``<base>/charts/``）
        html_dir: HTML 报告落盘目录；缺省 = output_dir。设为 ``base_dir`` 时
            HTML 落到上层目录、PNG 落到 charts/，与 compare 风格统一。
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
    except ImportError as e:
        logger.warning(f"matplotlib 不可用，跳过 PNG 渲染: {e}")
        return {}, None

    # 优先找系统装的 CJK 字体，避免中文标题渲染成 □
    cjk_candidates = [
        "PingFang SC", "Heiti SC", "STHeiti", "Hiragino Sans GB",
        "Arial Unicode MS", "Microsoft YaHei", "SimHei",
        "Noto Sans CJK SC", "Source Han Sans SC", "WenQuanYi Zen Hei",
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    chosen = next((c for c in cjk_candidates if c in available), None)
    if chosen:
        matplotlib.rcParams["font.sans-serif"] = [chosen] + list(
            matplotlib.rcParams.get("font.sans-serif", [])
        )
        matplotlib.rcParams["axes.unicode_minus"] = False
    else:
        logger.warning(
            "未找到 CJK 字体，PNG 中文将显示为 □；"
            "考虑安装 PingFang SC / Noto Sans CJK SC 后重跑"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    chart_paths: dict[str, Path] = {}

    # 1) Bar: mean ± std。指标分两组绘制：
    #   - 0/1 指标（ECR/TS/IFS/IISR/AR）公用 [0,1] 轴
    #   - Eff / CEI 不在 [0,1]，单独的 unbounded bar 图
    bounded_objs = [
        report.ecr, report.ts, report.ifs, report.iisr, report.ar,
    ]
    bounded_means = [m.mean for m in bounded_objs]
    bounded_stds = [m.std for m in bounded_objs]
    bounded_names = [m.metric for m in bounded_objs]

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(bounded_names, bounded_means, yerr=bounded_stds, capsize=6,
           color="#4472C4", alpha=0.85)
    ax.set_ylim(0, 1.05)
    ax.set_ylabel("mean")
    ax.set_title("0/1 指标 mean ± std")
    for i, m in enumerate(bounded_means):
        ax.text(i, m + 0.02, f"{m:.3f}", ha="center", fontsize=9)
    fig.tight_layout()
    p = output_dir / "metrics_bar.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    chart_paths["metrics_bar"] = p

    # 2) Box plot: per-metric distribution（限定 0/1 指标，统一 [0,1] 轴更直观）
    if rows:
        per_metric = [
            [float(r.metrics.get(m, 0.0) or 0.0) for r in rows] for m in bounded_names
        ]
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.boxplot(per_metric, labels=bounded_names, showmeans=True)
        ax.set_ylim(-0.05, 1.05)
        ax.set_ylabel("score")
        ax.set_title("0/1 指标 per-case 分布（箱线图）")
        fig.tight_layout()
        p = output_dir / "metrics_box.png"
        fig.savefig(p, dpi=140)
        plt.close(fig)
        chart_paths["metrics_box"] = p

    # 3) E2E_Latency histogram
    if report.e2e_latency_ms and rows:
        ttft_vals = [r.e2e_latency_ms for r in rows if r.e2e_latency_ms > 0]
        if ttft_vals:
            r = report.e2e_latency_ms
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.hist(ttft_vals, bins=20, color="#ED7D31", alpha=0.85)
            for label, val, color in (
                ("p50", r.p50, "#70AD47"),
                ("p95", r.p95, "#FFC000"),
                ("p99", r.p99, "#C00000"),
            ):
                ax.axvline(val, color=color, linestyle="--", linewidth=1.2,
                           label=f"{label}={val:.0f}ms")
            ax.set_xlabel("E2E Latency (ms)")
            ax.set_ylabel("count")
            ax.set_title(f"E2E_Latency 分布  N={r.count}  mean={r.mean:.1f}ms")
            ax.legend()
            fig.tight_layout()
            p = output_dir / "ttft_hist.png"
            fig.savefig(p, dpi=140)
            plt.close(fig)
            chart_paths["ttft_hist"] = p

    # 4) Tokens histogram
    if report.tokens_per_task and rows:
        token_vals = [r.total_tokens for r in rows if r.total_tokens > 0]
        if token_vals:
            r = report.tokens_per_task
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.hist(token_vals, bins=20, color="#5B9BD5", alpha=0.85)
            for label, val, color in (
                ("p50", r.p50, "#70AD47"),
                ("p95", r.p95, "#FFC000"),
            ):
                ax.axvline(val, color=color, linestyle="--", linewidth=1.2,
                           label=f"{label}={val:.0f}")
            ax.set_xlabel("total tokens / task")
            ax.set_ylabel("count")
            ax.set_title(f"Tokens 分布  N={r.count}  mean={r.mean:.1f}  sum={r.sum:.0f}")
            ax.legend()
            fig.tight_layout()
            p = output_dir / "tokens_hist.png"
            fig.savefig(p, dpi=140)
            plt.close(fig)
            chart_paths["tokens_hist"] = p

    # 5) HTML 一页式报告（嵌入 PNG + BatchReport 表）
    # html_dir 默认 = output_dir；外层指定时 HTML 落上层、PNG 仍在 output_dir
    target_html_dir = html_dir if html_dir is not None else output_dir
    target_html_dir.mkdir(parents=True, exist_ok=True)
    html_path = target_html_dir / "report.html"
    _write_html_report(html_path, report, counts, chart_paths, html_dir=target_html_dir)
    return chart_paths, html_path


def _write_html_report(
    html_path: Path,
    report: BatchReport,
    counts: StatusCounts,
    chart_paths: dict[str, Path],
    html_dir: Optional[Path] = None,
) -> None:
    """渲染单文件 HTML，图片用相对路径嵌入。

    HTML 与 PNG 可能不在同一目录（比如 HTML 在 base_dir、PNG 在 charts/），
    所以图片 src 用 ``Path.relative_to(html_dir)`` 计算相对路径。
    """
    base = html_dir if html_dir is not None else html_path.parent

    def _rel(p: Path) -> str:
        try:
            return str(p.relative_to(base))
        except ValueError:
            return p.name  # 不同盘符等情况兜底

    rows_html = []
    for m in (
        report.ecr, report.ts, report.ifs, report.iisr,
        report.ar, report.eff, report.cei,
    ):
        rows_html.append(
            f"<tr><td>{m.metric}</td><td>{m.n}</td><td>{m.mean:.4f}</td>"
            f"<td>{m.median:.4f}</td><td>{m.std:.4f}</td>"
            f"<td>{m.min:.4f}</td><td>{m.max:.4f}</td></tr>"
        )

    rt_lines = []
    if report.e2e_latency_ms:
        r = report.e2e_latency_ms
        rt_lines.append(
            f"<p><b>E2E_Latency(ms)</b>: N={r.count}, mean={r.mean:.1f}, p50={r.p50:.1f}, "
            f"p95={r.p95:.1f}, p99={r.p99:.1f}, max={r.max:.1f}</p>"
        )
    if report.tokens_per_task:
        r = report.tokens_per_task
        rt_lines.append(
            f"<p><b>Tokens/task</b>: N={r.count}, mean={r.mean:.1f}, "
            f"p50={r.p50:.1f}, p95={r.p95:.1f}, sum={r.sum:.0f}</p>"
        )

    img_tags = []
    for key, label in (
        ("metrics_bar", "0/1 指标 mean ± std"),
        ("metrics_box", "0/1 指标 per-case 分布（箱线）"),
        ("ttft_hist", "E2E_Latency 分布"),
        ("tokens_hist", "Tokens 分布"),
    ):
        if key in chart_paths:
            img_tags.append(
                f"<div class='chart'><h3>{label}</h3>"
                f"<img src='{_rel(chart_paths[key])}' /></div>"
            )

    html = f"""<!doctype html>
<html><head><meta charset='utf-8'><title>评测批次报告</title>
<style>
body{{font-family:-apple-system,Segoe UI,sans-serif;margin:2em;color:#222;}}
h1{{border-bottom:2px solid #4472C4;padding-bottom:6px;}}
table{{border-collapse:collapse;margin:1em 0;}}
th,td{{border:1px solid #ccc;padding:6px 12px;text-align:right;}}
th{{background:#4472C4;color:white;}}
td:first-child{{text-align:left;font-weight:bold;}}
.chart{{margin:1.5em 0;}}
.chart img{{max-width:900px;border:1px solid #eee;border-radius:4px;}}
.summary{{background:#f5f5f5;padding:1em;border-radius:4px;}}
</style></head><body>
<h1>评测批次报告</h1>
<div class='summary'>
  <p>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
  <p>样本：total={counts.n_total} | success={counts.n_success} |
     skipped={counts.n_skipped} | error={counts.n_error} |
     unparseable={counts.n_unparseable}</p>
</div>
<h2>评估指标统计</h2>
<table>
<tr><th>metric</th><th>n</th><th>mean</th><th>median</th>
    <th>std</th><th>min</th><th>max</th></tr>
{''.join(rows_html)}
</table>
<h2>Runtime</h2>
{''.join(rt_lines) if rt_lines else '<p><i>(无 runtime 数据)</i></p>'}
<h2>可视化</h2>
{''.join(img_tags) if img_tags else '<p><i>(无图表)</i></p>'}
</body></html>"""
    html_path.write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------


def summarize(
    xlsx_path: Path,
    output_dir: Path,
    *,
    top_n: int = 5,
    visualize: VisualizeMode = "both",
) -> SummaryProducts:
    """读 xlsx → 聚合 → 落盘所有产物。这是 HTTP 路由 / CLI 共用入口。

    产物路径布局（每次任务的所有文件聚合到一个时间戳目录）：

        <output_dir>/single/<ts>/
        ├── summary.json
        ├── summary.txt
        ├── summary.xlsx
        ├── report.html
        └── charts/
            ├── metrics_bar.png
            └── ...

    Args:
        xlsx_path: ``data/outputs/evaluation_res/`` 下的评测结果 xlsx
        output_dir: 产物落盘根目录（建议 ``data/outputs/report``）
        top_n: outlier 列每个指标各 top-N 高/低
        visualize: ``xlsx`` / ``png`` / ``both`` / ``none``

    Returns:
        SummaryProducts 包含所有生成文件的路径
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    rows, counts = load_eval_results(xlsx_path)
    logger.info(
        f"读入 {counts.n_total} 行 → success={counts.n_success} "
        f"skipped={counts.n_skipped} error={counts.n_error} "
        f"unparseable={counts.n_unparseable}"
    )

    # 7 指标 + CEI 的 scores 列表用全部行（失败 case 已被 _zero_eval_row 置 0），
    # runtime stats（E2E_Latency / tokens）只看 success 行避免被 0 拉低 p95。
    scores_list = build_metric_scores(rows)
    success_rows = [r for r in rows if r.status == "success"]
    ttft_per_case = [r.e2e_latency_ms for r in success_rows]
    tokens_per_case = [r.total_tokens for r in success_rows]
    # CEI：与 0/1 指标同口径，全量 case 都纳入分母（失败/缺测按 0），
    # 走独立 list 而非 MetricScores（CEI 不在 judge schema 里、由 add_cei.py 后写）
    cei_per_case = [float(r.metrics.get("CEI", 0.0) or 0.0) for r in rows]
    report = aggregate_batch(
        scores_list,
        e2e_latency_ms_per_case=ttft_per_case,
        tokens_per_case=tokens_per_case,
        cei_per_case=cei_per_case,
    )

    outliers = find_outliers(rows, top_n=top_n)

    # 控制台简报
    print(format_batch_report(report))

    # 单文件统计任务的所有产物聚合到 <output_dir>/single/<ts>/
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_dir = output_dir / "single" / ts
    base_dir.mkdir(parents=True, exist_ok=True)

    json_path = base_dir / "summary.json"
    txt_path = base_dir / "summary.txt"
    render_json(report, counts, outliers, json_path)
    render_txt(report, counts, txt_path)

    products = SummaryProducts(json_path=json_path, txt_path=txt_path)

    if visualize in ("xlsx", "both"):
        xlsx_out = base_dir / "summary.xlsx"
        render_xlsx(report, rows, outliers, counts, xlsx_out)
        products.xlsx_path = xlsx_out

    if visualize in ("png", "both"):
        charts_dir = base_dir / "charts"
        chart_paths, html_path = render_png(
            report, rows, counts, charts_dir, html_dir=base_dir
        )
        products.charts_dir = charts_dir if chart_paths else None
        products.chart_paths = chart_paths
        products.html_path = html_path

    return products
