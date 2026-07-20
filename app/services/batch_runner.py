import asyncio
import logging
import time
import uuid
from pathlib import Path
from typing import Optional

from langchain_core.messages import HumanMessage, SystemMessage
from openpyxl import load_workbook

from app.core.simulator.agent_simulator import BaseSimulationAgent
from app.schemas.batch_evaluate import EvaluateCase, EvaluateResult

logger = logging.getLogger(__name__)


class BatchResultWriter:
    """批量评测的缓冲式 / 并发安全 xlsx 追加写入器。

    - load_workbook 一次进内存，每 ``flush_every`` 条触发一次 ``wb.save``，
      避免每条 case 重写整文件的 IO 风暴。
    - 所有 append/flush 走 ``asyncio.Lock`` 串行化，多并发 case 写同一个文件
      不会撕裂 xlsx。
    - ``save`` 走 ``asyncio.to_thread`` 不阻塞 event loop。

    用法：
        path = create_batch_output_file(input_filename)   # 先建带表头的空文件
        async with BatchResultWriter(path) as writer:
            await writer.append(case, result, case_index=1)
            ...
        # 退出时 close → flush
    """

    def __init__(self, path: Path, flush_every: int = 5):
        self.path = Path(path)
        self.flush_every = max(1, flush_every)
        self._wb = load_workbook(self.path)
        self._ws = self._wb.active
        self._unflushed = 0
        self._lock = asyncio.Lock()
        self._closed = False

    @staticmethod
    def _row_for(case: EvaluateCase, result: EvaluateResult, case_index: int) -> list:
        return [
            case_index,
            case.query,
            case.location or "",
            case.tool or "",
            case.expected or "",
            result.conversation_id,
            result.status,
            result.final_response,
            str(result.tool_calls) if result.tool_calls else "",
            result.execution_time_ms,
            result.error_message or "",
        ]

    async def append(
        self,
        case: EvaluateCase,
        result: EvaluateResult,
        case_index: int,
    ) -> None:
        async with self._lock:
            if self._closed:
                logger.warning(
                    "BatchResultWriter 已关闭，丢弃 result (case_index=%s, conv_id=%s)",
                    case_index, result.conversation_id,
                )
                return
            self._ws.append(self._row_for(case, result, case_index))
            self._unflushed += 1
            if self._unflushed >= self.flush_every:
                await self._flush_locked()

    async def flush(self) -> None:
        async with self._lock:
            if not self._closed:
                await self._flush_locked()

    async def close(self) -> None:
        async with self._lock:
            if self._closed:
                return
            await self._flush_locked()
            self._closed = True

    async def _flush_locked(self) -> None:
        """调用方需持有 self._lock。save 失败不重置 unflushed，下次会重试。"""
        if self._unflushed == 0:
            return
        try:
            await asyncio.to_thread(self._wb.save, str(self.path))
            self._unflushed = 0
        except Exception as e:
            logger.warning("BatchResultWriter 保存失败 (path=%s): %s", self.path, e)

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        await self.close()


class BatchRunner:
    """批量评测执行引擎。"""

    def __init__(self, agent: BaseSimulationAgent):
        self.agent = agent

    async def run_case(
        self,
        case: EvaluateCase,
        thread_id: str,
    ) -> EvaluateResult:
        """单条用例执行（case 内串行：单次 graph.ainvoke 跑完整 React loop）。"""
        conversation_id = str(uuid.uuid4())
        start_time = time.time()

        try:
            messages = []
            if case.location:
                messages.append(
                    SystemMessage(content=f"当前位置：{case.location}")
                )
            messages.append(HumanMessage(content=case.query))

            state = {"messages": messages}
            config = {"configurable": {"thread_id": thread_id}}

            graph = self.agent.run_graph()
            result_state = await asyncio.wait_for(
                graph.ainvoke(state, config),
                timeout=60.0,
            )

            final_messages = result_state.get("messages", [])
            final_response = ""
            tool_calls: list[dict] = []

            # 先收集所有 ToolMessage 的响应，按 tool_call_id 映射
            tool_responses = {}
            for msg in final_messages:
                msg_type = getattr(msg, "type", "")
                if msg_type == "tool":
                    tid = getattr(msg, "tool_call_id", "")
                    tool_responses[tid] = str(getattr(msg, "content", "") or "")

            # 只提取**最后一条** AI 消息的 tool_calls（当前轮次的调用）
            for msg in reversed(final_messages):
                if hasattr(msg, "content") and getattr(msg, "type", "") == "ai":
                    # ChatAnthropic content 可能是 list[blocks]（thinking on
                    # 时含 thinking signature），直接 str() 会把整个结构
                    # （含 ~3KB signature）写进 EvaluateResult。走 normalize
                    # 保证只取 text 块。
                    final_response = (
                        BaseSimulationAgent._normalize_anthropic_content_to_str(
                            msg.content
                        )
                    )
                    if hasattr(msg, "tool_calls") and msg.tool_calls:
                        for tc in msg.tool_calls:
                            tool_calls.append({
                                "name": tc.get("name", ""),
                                "args": tc.get("args", {}),
                                "id": tc.get("id", ""),
                                "response": tool_responses.get(tc.get("id", ""), ""),
                            })
                    # 找到最后一条 AI 消息后立即退出
                    break

            execution_time_ms = int((time.time() - start_time) * 1000)

            return EvaluateResult(
                conversation_id=conversation_id,
                query=case.query,
                location=case.location,
                tool=case.tool,
                expected=case.expected,
                status="success",
                final_response=final_response,
                tool_calls=tool_calls,
                execution_time_ms=execution_time_ms,
                error_message=None,
            )

        except asyncio.TimeoutError:
            execution_time_ms = int((time.time() - start_time) * 1000)
            return EvaluateResult(
                conversation_id=conversation_id,
                query=case.query,
                location=case.location,
                tool=case.tool,
                expected=case.expected,
                status="timeout",
                final_response="",
                tool_calls=[],
                execution_time_ms=execution_time_ms,
                error_message="执行超时（60s）",
            )

        except Exception as e:
            execution_time_ms = int((time.time() - start_time) * 1000)
            logger.exception(f"用例执行失败: {case.query}")
            return EvaluateResult(
                conversation_id=conversation_id,
                query=case.query,
                location=case.location,
                tool=case.tool,
                expected=case.expected,
                status="error",
                final_response="",
                tool_calls=[],
                execution_time_ms=execution_time_ms,
                error_message=str(e),
            )

    async def run_batch(
        self,
        cases: list[EvaluateCase],
        thread_id_prefix: str | None = None,
        writer: Optional[BatchResultWriter] = None,
        max_concurrency: int = 4,
    ) -> list[EvaluateResult]:
        """批量执行：case 间并行（受 max_concurrency 限制），case 内串行。

        Args:
            cases: 用例列表
            thread_id_prefix: thread_id 前缀；并发安全（每条 case 用 idx 拼唯一 thread_id）
            writer: 可选的流式写入器；传入则每条 case 完成后立即追加写盘
            max_concurrency: 并发上限（默认 4）

        Returns:
            按输入顺序对齐的结果列表（asyncio.gather 保留顺序，与完成时间无关）
        """
        prefix = thread_id_prefix or "eval"
        n = len(cases)
        sem = asyncio.Semaphore(max(1, max_concurrency))

        async def _run_one(idx: int, case: EvaluateCase) -> EvaluateResult:
            async with sem:
                thread_id = f"{prefix}_{idx + 1}"
                preview = (case.query or "")[:40]
                logger.info(f"[{idx + 1}/{n}] start | query={preview}...")
                result = await self.run_case(case, thread_id=thread_id)
                logger.info(
                    f"[{idx + 1}/{n}] {result.status} | "
                    f"conv_id={result.conversation_id} | query={preview}..."
                )
                if writer is not None:
                    try:
                        await writer.append(case, result, case_index=idx + 1)
                    except Exception as e:
                        logger.warning(
                            f"[{idx + 1}] writer.append 失败（不影响主流程）: {e}"
                        )
                return result

        # gather 返回值与输入顺序对齐；并发实际完成顺序可能不同，但
        # results[i] 永远对应 cases[i]
        return await asyncio.gather(
            *(_run_one(i, c) for i, c in enumerate(cases))
        )
