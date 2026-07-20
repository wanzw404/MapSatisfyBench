"""从 simulator_res Excel 批量读取对话结果并调用评测服务。

Usage:
    cd <repo-root>
    .venv/bin/python -m app.scripts.batch_evaluate_from_simulator \
        --input data/outputs/simulator_res/dialogue_xxx.xlsx \
        --cases-input data/inputs/parsed_demo.xlsx \
        --output data/outputs/evaluation_res \
        --max-concurrency 2

聚合规则：
  - 同 conversation_id 内按 turn_index 升序、同 turn_index 内 user 在前 assistant 在后排序
  - 重新编号 turn_index 使其从 1 开始：
      role=user      →  output_turn_index = 当前 turn_index + 1
      role=assistant →  output_turn_index = 当前 turn_index
    （初始 user 0→1；第 t 轮 agent t→t；第 t 轮 user t→t+1）
  - 跳过 is_forced_stop=true 与 status != success 的行
  - tool_calls 解析失败的 turn 仍保留消息，但记入 parse_errors 供审计
  - ground_truth 解析失败 / 缺失时 case 标 status=skipped
  - full_intent 通过 --cases-input 文件按 task_id=conversation_id 关联取得
"""

import argparse
import ast
import asyncio
import csv
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import HTTPException
from openpyxl import load_workbook  # 仅用于兼容老 xlsx 输入；评测结果改写 CSV

# 项目根加入 sys.path 便于直接 python -m 运行
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.schemas.evaluate_schemas import (  # noqa: E402
    ConversationMessage,
    EvaluateRequest,
    SessionStats,
)
from app.services.evaluate_service import run_evaluation  # noqa: E402

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")


# ─────────────────────────────────────────────────────────────────────
# Excel 读取
# ─────────────────────────────────────────────────────────────────────

def read_simulator_results(file_path: Path) -> list[dict]:
    """读 simulator_res 文件 → list[dict]。

    自动按扩展名分流：
      - ``.csv``  → 标准 csv 模块（utf-8-sig 自动剥 BOM）；新版 dialogue_simulator 输出
      - ``.xlsx`` → openpyxl（保留向后兼容，老文件仍能读）
    """
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(file_path)
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(file_path)
    raise ValueError(
        f"不支持的输入格式: {file_path.suffix} (期望 .csv 或 .xlsx)"
    )


def _read_csv(file_path: Path) -> list[dict]:
    # Python csv 默认单字段上限 128KB；dialogue 仿真 CSV 的 tool_calls 列
    # 经常含多条长 tool 响应拼在一起（每条最大 16KB × N 条），轻松超过。
    # 提到 sys.maxsize 让 csv 不再卡字段长度——文件大小本身仍受 OS 限制。
    csv.field_size_limit(sys.maxsize)
    rows: list[dict] = []
    with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # csv.DictReader 返回的全是 str；空格 cell → "" 而非 None
            # 下游 int(... or 0) / str(...) 兼容
            if not row or all(
                v is None or str(v).strip() == "" for v in row.values()
            ):
                continue
            rows.append(dict(row))
    return rows


def _read_xlsx(file_path: Path) -> list[dict]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


# 别名：保留旧函数名供已有调用方使用
read_simulator_excel = read_simulator_results


def load_full_intent_map(cases_path: Optional[Path]) -> dict[str, str]:
    """从原始 input xlsx 读取 task_id → full_intent 映射。"""
    if not cases_path:
        logger.info("未指定 --cases-input；full_intent 将统一留空")
        return {}
    if not cases_path.exists():
        logger.warning(f"--cases-input 文件不存在: {cases_path}；full_intent 将统一留空")
        return {}

    wb = load_workbook(cases_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    if "task_id" not in headers:
        logger.warning(f"{cases_path} 未找到 task_id 列；full_intent 将统一留空")
        return {}
    if "full_intent" not in headers:
        logger.warning(f"{cases_path} 未找到 full_intent 列；full_intent 将统一留空")
        return {}
    tid_idx = headers.index("task_id")
    fi_idx = headers.index("full_intent")

    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(tid_idx, fi_idx):
            continue
        tid = str(row[tid_idx]).strip() if row[tid_idx] is not None else ""
        fi = str(row[fi_idx]).strip() if row[fi_idx] is not None else ""
        if tid:
            mapping[tid] = fi
    logger.info(f"从 {cases_path} 读取 {len(mapping)} 条 task_id→full_intent 映射")
    return mapping


# ─────────────────────────────────────────────────────────────────────
# Cell 解析
# ─────────────────────────────────────────────────────────────────────

def parse_tool_calls(cell_value) -> Optional[list]:
    """解析 tool_calls cell。

    Returns:
        - list: 解析成功（可能是 []）
        - None: 解析失败（cell 截断 / 格式损坏 / 非 list）。调用方据此打 parse_errors 标记。
    """
    if cell_value is None or str(cell_value).strip() == "":
        return []
    raw = str(cell_value).strip()

    # 1) JSON 优先（新写出端用 json.dumps；当前是 str(list[dict])，会失败转下一档）
    try:
        result = json.loads(raw)
        if isinstance(result, list):
            return result
    except json.JSONDecodeError:
        pass

    # 2) Python repr 兜底（旧 xlsx 用 str(list[dict]) 写入）
    try:
        result = ast.literal_eval(raw)
        if isinstance(result, list):
            return result
    except (ValueError, SyntaxError):
        pass

    # 3) 截断兜底：找最后一个 } 截断 + 补 ]
    last_brace = raw.rfind("}")
    if last_brace > 0:
        candidate = raw[: last_brace + 1] + "]"
        for parser in (json.loads, ast.literal_eval):
            try:
                result = parser(candidate)
                if isinstance(result, list):
                    return result
            except Exception:
                pass

    logger.warning(
        f"无法解析 tool_calls (len={len(raw)}, head={raw[:120]!r})"
    )
    return None


def parse_ground_truth(cell_value) -> Optional[dict]:
    """解析 ground_truth cell。非字典或解析失败返回 None。

    ground_truth 列实际可能是「业务包装结构」：
        {task_id, meta_info, input_data, user_simulator, ground_truth: {...真rubric...}}
    而 ``GroundTruth`` schema 期望直接拿到 rubric（含 explicit_intent /
    implicit_intent / truth_trajectory）。如果直接把外层 dict 喂给 model_validate，
    因为 ``extra="allow"`` 不会报错，但 ``gt.truth_trajectory`` 等字段会全部
    fallback 到默认空值——直接访问该字段的 metric（如 TS）会归零。

    本函数检测到这种包装层时，下钻一层 ``.ground_truth`` 取真实 rubric。
    """
    if cell_value is None or str(cell_value).strip() == "":
        return None
    raw = str(cell_value).strip()
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return None
    if not isinstance(parsed, dict):
        return None

    # 判定 + unwrap：顶层若不含 rubric 标志字段、且嵌套的 .ground_truth 子字段
    # 含有这些字段，就取里层。
    rubric_keys = {"explicit_intent", "implicit_intent", "truth_trajectory"}
    has_rubric_top = any(k in parsed for k in rubric_keys)
    inner = parsed.get("ground_truth")
    if not has_rubric_top and isinstance(inner, dict) and any(
        k in inner for k in rubric_keys
    ):
        logger.info(
            "ground_truth 是业务包装层，已下钻 .ground_truth 取真实 rubric"
        )
        return inner
    return parsed


def _is_forced_stop(row: dict) -> bool:
    return str(row.get("is_forced_stop", "")).strip().lower() == "true"


def _status_ok(row: dict) -> bool:
    s = str(row.get("status", "success")).strip().lower()
    return s in ("success", "")


def _extract_time_and_location_from_context(
    group_rows: list[dict],
) -> tuple[str, str]:
    """从 group 内任一行的 ``context`` JSON 提取 ``time`` / ``user_loc_name``，
    送给评分模块的 ``current_time`` / ``current_location``。

    优先级：
      1. ``context`` JSON（仿真器真实接收的上下文，最权威）
      2. ``time`` / ``location`` 列兜底（来自 DialogueCase 入参；老 CSV 没
         context 列时仍可工作）

    扫每一行，已填满任一字段就不再覆盖；两个字段都填满立即停止。
    context JSON 解析失败 → 跳过本行的 context 字段（DEBUG 日志），走 columns
    兜底，不阻塞主流程。
    """
    current_time = ""
    current_location = ""
    for r in group_rows:
        # 1) 优先解析 context JSON
        ctx_raw = r.get("context")
        if ctx_raw and str(ctx_raw).strip():
            try:
                ctx = json.loads(str(ctx_raw))
                if isinstance(ctx, dict):
                    if not current_time and ctx.get("time"):
                        current_time = str(ctx["time"]).strip()
                    if not current_location and ctx.get("user_loc_name"):
                        current_location = str(ctx["user_loc_name"]).strip()
            except (json.JSONDecodeError, TypeError) as e:
                logger.debug(
                    "[%s] context JSON 解析失败，跳过本行 context 字段: %s",
                    r.get("conversation_id", "<?>"), e,
                )
        # 2) fallback 到 time / location 列
        if not current_time and r.get("time"):
            current_time = str(r.get("time")).strip()
        if not current_location and r.get("location"):
            current_location = str(r.get("location")).strip()
        if current_time and current_location:
            break
    return current_time, current_location


# ─────────────────────────────────────────────────────────────────────
# 聚合
# ─────────────────────────────────────────────────────────────────────

def aggregate_by_conversation(
    rows: list[dict],
    full_intent_map: dict[str, str],
) -> dict[str, dict]:
    """按 conversation_id 聚合并按 EvaluateRequest 协议组装 case 数据。"""
    groups: dict[str, list[dict]] = {}
    for row in rows:
        cid = str(row.get("conversation_id", "")).strip()
        if not cid:
            continue
        groups.setdefault(cid, []).append(row)

    aggregated: dict[str, dict] = {}
    for cid, group_rows in groups.items():
        # 1) 过滤强制截断 / 异常行 + 仅保留 user/assistant
        clean_rows = [
            r
            for r in group_rows
            if not _is_forced_stop(r)
            and _status_ok(r)
            and str(r.get("role", "")).strip() in ("user", "assistant")
        ]
        if not clean_rows:
            logger.warning(f"[{cid}] 过滤后无有效消息，跳过")
            continue

        # 2) 重编号 turn_index
        for r in clean_rows:
            cur = int(r.get("turn_index", 0) or 0)
            r["_output_turn_index"] = (
                cur + 1 if r.get("role") == "user" else cur
            )

        # 3) 排序：先按重编号后的 turn_index 升序，同 turn_index 内 user 在前
        clean_rows.sort(
            key=lambda r: (
                r["_output_turn_index"],
                0 if r.get("role") == "user" else 1,
            )
        )

        # 4) 构造 conversation_history_messages + 累计 tokens
        messages: list[dict] = []
        total_input_tokens = 0
        total_output_tokens = 0
        parse_errors: list[int] = []

        for r in clean_rows:
            role = str(r.get("role", "")).strip()
            content = str(r.get("content", "") or "")
            tidx = int(r["_output_turn_index"])
            execution_time_ms = int(r.get("execution_time_ms", 0) or 0)
            # 新版 csv 用 input_tokens / output_tokens 两列；旧版 xlsx 是单列 content_tokens
            # （user 行=output、assistant 行=output，input 维度从未记），向后兼容 fallback
            content_tokens_legacy = int(r.get("content_tokens", 0) or 0)
            input_tokens_row = int(
                r.get("input_tokens", 0) or 0
            ) or content_tokens_legacy
            output_tokens_row = int(
                r.get("output_tokens", 0) or 0
            ) or content_tokens_legacy

            if role == "user":
                total_input_tokens += input_tokens_row
                messages.append(
                    {
                        "role": "user",
                        "content": content,
                        "turn_index": tidx,
                    }
                )
            else:  # assistant
                total_output_tokens += output_tokens_row
                tcs = parse_tool_calls(r.get("tool_calls"))
                if tcs is None:
                    parse_errors.append(tidx)
                    tcs = []
                messages.append(
                    {
                        "role": "assistant",
                        "content": content,
                        "turn_index": tidx,
                        "tool_calls": tcs,
                        "TTFT": execution_time_ms,
                    }
                )

        # 5) case 级字段
        query = next(
            (m["content"] for m in messages if m["role"] == "user"),
            "",
        )

        ground_truth: Optional[dict] = None
        for r in group_rows:  # 找 group 内任一行的非空 ground_truth
            gt = parse_ground_truth(r.get("ground_truth"))
            if gt:
                ground_truth = gt
                break

        # context.time / context.user_loc_name 优先；time/location 列兜底
        current_time, current_location = _extract_time_and_location_from_context(
            group_rows
        )

        full_intent = full_intent_map.get(cid, "")

        aggregated[cid] = {
            "case_id": cid,
            "query": query,
            "full_intent": full_intent,
            "persona": "",  # 用户明确不关注
            "current_time": current_time,
            "current_location": current_location,
            "conversation_history_messages": messages,
            "ground_truth": ground_truth,
            "session_stats": {
                "total_input_tokens": total_input_tokens,
                "total_output_tokens": total_output_tokens,
            },
            # 元数据：审计用，不影响 EvaluateRequest 构造
            "_meta": {
                "parse_errors": parse_errors,  # tool_calls 解析失败的 turn_index 列表
                "n_messages": len(messages),
            },
        }

    return aggregated


# ─────────────────────────────────────────────────────────────────────
# 单 case 评测
# ─────────────────────────────────────────────────────────────────────

async def evaluate_one(cid: str, data: dict, opts: dict) -> dict:
    """对单 conversation 调评测。任何异常都被吞为 status=error 不外抛。"""
    parse_errors = data.get("_meta", {}).get("parse_errors", [])

    # ground_truth 缺失时跳过——空 dict 喂下游会让 ECR/IISR 全归零
    if data.get("ground_truth") is None:
        logger.warning(f"[{cid}] ground_truth 缺失或非合法 JSON，跳过评测")
        return {
            "case_id": cid,
            "results": {},
            "reason": "ground_truth missing or invalid",
            "status": "skipped",
            "error": "missing or invalid ground_truth",
            "parse_errors": parse_errors,
        }

    try:
        conversation_history = [
            ConversationMessage.model_validate(msg)
            for msg in data["conversation_history_messages"]
        ]
        req = EvaluateRequest(
            case_id=cid,
            query=data["query"],
            full_intent=data.get("full_intent", "") or "",
            persona=data.get("persona", "") or "",
            current_time=data.get("current_time", "") or "",
            current_location=data.get("current_location", "") or "",
            conversation_history_messages=conversation_history,
            ground_truth=data["ground_truth"],
            session_stats=SessionStats(
                total_input_tokens=data["session_stats"]["total_input_tokens"],
                total_output_tokens=data["session_stats"]["total_output_tokens"],
            ),
            language=opts.get("language", "chinese"),
            enable_verification=opts.get("enable_verification", True),
            enable_meta_judge=opts.get("enable_meta_judge", False),
            model=opts.get("model", "") or "",
        )

        resp = await run_evaluation(req)
        return {
            "case_id": resp.case_id,
            "results": resp.results,
            "reason": resp.reason,
            "status": "success",
            "error": "",
            "parse_errors": parse_errors,
        }
    except HTTPException as he:
        # 区分 400 (gt 校验)、502 (LLM 非 JSON)、500 (其它) 等
        logger.exception(f"[{cid}] 评测 HTTPException ({he.status_code})")
        return {
            "case_id": cid,
            "results": {},
            "reason": "",
            "status": "error",
            "error": f"HTTP {he.status_code}: {he.detail}",
            "parse_errors": parse_errors,
        }
    except Exception as exc:
        logger.exception(f"[{cid}] 评测失败")
        return {
            "case_id": cid,
            "results": {},
            "reason": "",
            "status": "error",
            "error": f"{type(exc).__name__}: {exc}",
            "parse_errors": parse_errors,
        }


# ─────────────────────────────────────────────────────────────────────
# 输出 / 编排
# ─────────────────────────────────────────────────────────────────────

EVAL_OUTPUT_HEADERS = [
    "case_id",
    "status",
    "error",
    "parse_errors",
    "results",
    "reason",
]


def _record_to_row(r: dict) -> list:
    """单条评测结果 → CSV 行（与 EVAL_OUTPUT_HEADERS 对齐）。"""
    return [
        r.get("case_id", ""),
        r.get("status", ""),
        r.get("error", "") or "",
        json.dumps(r.get("parse_errors") or [], ensure_ascii=False),
        json.dumps(r["results"], ensure_ascii=False) if r.get("results") else "",
        r.get("reason", "") or "",
    ]


def create_eval_output_file(
    output_dir: Path,
    ts: str,
    suffix_part: str = "",
) -> Path:
    """提前建 ``evaluation_result_<ts>[_<suffix>].csv`` 并写入表头。

    与 dialogue_recorder.create_output_file 同模式：utf-8-sig BOM 让 Excel
    双击打开能识别中文；后续由 EvaluationResultWriter 在 append 模式下追加。
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"evaluation_result_{ts}{suffix_part}.csv"
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(EVAL_OUTPUT_HEADERS)
    logger.info(f"评测结果文件已创建（含表头）: {output_path}")
    return output_path


class EvaluationResultWriter:
    """缓冲式、并发安全的 evaluation_result_*.csv 追加写入器。

    与 ``app.services.dialogue_recorder.DialogueResultWriter`` 同模式：
      - 持久 file handle + ``csv.writer``，避免每条 append 都打开关闭文件
      - ``asyncio.Lock`` 串行化 ``writerow``，并发 case 完成时不会撕裂
      - ``flush_every`` 控制把 buffer 推到内核（``fh.flush``）的频率
      - ``utf-8-sig``：BOM + UTF-8，Excel/WPS 双击能正确识别中文

    用法：
        async with EvaluationResultWriter(path) as writer:
            await writer.append(record)
            ...
        # 退出时自动 close → flush
    """

    def __init__(self, path: Path, flush_every: int = 5):
        self.path = Path(path)
        self.flush_every = max(1, flush_every)
        # append 模式打开（表头已由 create_eval_output_file 写好）；不靠
        # atexit，明确通过 close()/__aexit__ 释放
        self._fh = open(self.path, "a", newline="", encoding="utf-8-sig")
        self._writer = csv.writer(self._fh, quoting=csv.QUOTE_MINIMAL)
        self._unflushed = 0
        self._lock = asyncio.Lock()
        self._closed = False

    async def append(self, record: dict) -> None:
        async with self._lock:
            if self._closed:
                logger.warning(
                    "EvaluationResultWriter 已关闭，丢弃 record (case_id=%s)",
                    record.get("case_id"),
                )
                return
            self._writer.writerow(_record_to_row(record))
            self._unflushed += 1
            if self._unflushed >= self.flush_every:
                await self._flush_locked()

    async def flush(self) -> None:
        async with self._lock:
            if not self._closed:
                await self._flush_locked()

    async def close(self) -> None:
        async with self._lock:
            if self._closed:
                return
            await self._flush_locked()
            try:
                self._fh.close()
            except Exception as e:
                logger.warning(
                    "EvaluationResultWriter 关闭文件句柄异常 (path=%s): %s",
                    self.path, e,
                )
            self._closed = True

    async def _flush_locked(self) -> None:
        """调用方需持有 self._lock；只 flush Python buffer 到内核，不强制 fsync。"""
        if self._unflushed == 0:
            return
        try:
            await asyncio.to_thread(self._fh.flush)
            self._unflushed = 0
        except Exception as e:
            logger.warning(
                "EvaluationResultWriter flush 失败 (path=%s): %s", self.path, e
            )

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        await self.close()


def _sanitize_filename_suffix(suffix: str) -> str:
    """与 dialogue_recorder._sanitize_filename_suffix 同语义；这里复刻一份避免循环依赖。"""
    import re as _re
    cleaned = _re.sub(r"[^A-Za-z0-9._-]+", "_", suffix or "").strip("_")
    return cleaned[:60]


async def batch_evaluate(
    input_path: Path,
    output_dir: Path,
    cases_input: Optional[Path],
    max_concurrency: int,
    opts: dict,
    suffix: Optional[str] = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    rows = read_simulator_excel(input_path)
    logger.info(f"从 {input_path} 读取 {len(rows)} 行原始数据")

    full_intent_map = load_full_intent_map(cases_input)

    aggregated = aggregate_by_conversation(rows, full_intent_map)
    logger.info(f"聚合为 {len(aggregated)} 个 conversation")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix_part = ""
    if suffix:
        cleaned = _sanitize_filename_suffix(suffix)
        if cleaned:
            suffix_part = f"_{cleaned}"

    # 1) 先把聚合结果落 JSON（评测前可审查）
    json_path = output_dir / f"aggregated_{ts}{suffix_part}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(aggregated, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"聚合 JSON 已写入: {json_path}")

    # 2) 提前建评测结果 CSV（写表头），后续每条 case 完成立即 append
    csv_path = create_eval_output_file(output_dir, ts, suffix_part)

    # 3) 并发评测 + 完成即写
    sem = asyncio.Semaphore(max(1, max_concurrency))
    total = len(aggregated)
    done_counter = 0

    async with EvaluationResultWriter(csv_path) as writer:
        async def _run(cid: str, data: dict) -> dict:
            nonlocal done_counter
            async with sem:
                logger.info(f"开始评测: {cid}")
                r = await evaluate_one(cid, data, opts)
            # 立即落盘——锁在 writer.append 内部，不要被 sem 包住，
            # 让其它 case 早点拿到 sem 进 LLM
            await writer.append(r)
            done_counter += 1
            logger.info(
                f"完成评测 [{done_counter}/{total}]: {cid} (status={r['status']}) → 已落盘"
            )
            return r

        results = await asyncio.gather(
            *[_run(c, d) for c, d in aggregated.items()]
        )

    logger.info(f"评测结果已全部写入: {csv_path}")

    # 4) 总览
    n_total = len(results)
    n_success = sum(1 for r in results if r["status"] == "success")
    n_skipped = sum(1 for r in results if r["status"] == "skipped")
    n_error = sum(1 for r in results if r["status"] == "error")
    n_parse_err = sum(1 for r in results if r.get("parse_errors"))
    logger.info(
        f"总览 → total={n_total} success={n_success} "
        f"skipped={n_skipped} error={n_error}  "
        f"(其中 tool_calls 解析失败的 case={n_parse_err})"
    )

    return csv_path


def main():
    parser = argparse.ArgumentParser(
        description="从 simulator_res 批量调用评测服务"
    )
    parser.add_argument(
        "--input",
        required=True,
        help="simulator_res Excel 路径（dialogue_simulator 写出的对话结果）",
    )
    parser.add_argument(
        "--cases-input",
        default=None,
        help=(
            "原始用例 xlsx（含 task_id + full_intent 列），用于按 task_id="
            "conversation_id 注入 full_intent。可选；不指定则 full_intent 留空。"
        ),
    )
    parser.add_argument(
        "--output",
        default="data/outputs/evaluation_res",
        help="输出目录（默认 data/outputs/evaluation_res）",
    )
    parser.add_argument(
        "--max-concurrency",
        type=int,
        default=2,
        help="LLM 并发上限（默认 2）",
    )
    parser.add_argument(
        "--language",
        default="chinese",
        choices=["chinese", "english"],
    )
    parser.add_argument(
        "--no-verify",
        action="store_true",
        help="关闭 web 事实校验（更快，不打 MCP gateway）",
    )
    parser.add_argument(
        "--enable-meta-judge",
        action="store_true",
        help="开启 Meta-Judge 复核（默认关闭）",
    )
    parser.add_argument(
        "--model",
        default="",
        help="覆盖默认 LLM 模型（默认走 settings.MODEL_NAME）",
    )
    parser.add_argument(
        "--suffix",
        default=None,
        help=(
            "输出 aggregated_*.json 与 evaluation_result_*.csv 文件名的后缀"
            "（在 timestamp 之后）。通常传 agent 模型名,用于矩阵评测时按模型"
            "区分产物。"
        ),
    )
    parser.add_argument(
        "--llm-qps",
        type=float,
        default=100.0,
        help=(
            "评分器侧 LLM 调用 QPS 上限（全局共享，所有 case / judge / "
            "verifier / meta_judge 共用一桶）。默认 100；<=0 关闭限流。"
            "另含 in-flight Semaphore=2*qps 兜底防止排队堆积，以及网络层"
            "错误自动重试 1 次（与 agent_simulator retry 口径对齐）。"
        ),
    )
    args = parser.parse_args()

    # 通过环境变量把 qps 透传给 evaluate_service._build_llm_provider，
    # 后者在构造 LLM provider 时按此值套上 RateLimitedRetryLLMProvider wrapper。
    # 走环境变量而非函数入参，避免在 service 层链路（HTTP / CLI 共用入口）
    # 上新增一个一直要透传到底的参数。
    os.environ["LLM_QPS"] = str(args.llm_qps)

    input_path = Path(args.input)
    output_dir = Path(args.output)
    cases_input = Path(args.cases_input) if args.cases_input else None

    if not input_path.exists():
        logger.error(f"输入文件不存在: {input_path}")
        sys.exit(1)

    opts = {
        "language": args.language,
        "enable_verification": not args.no_verify,
        "enable_meta_judge": args.enable_meta_judge,
        "model": args.model,
    }

    asyncio.run(
        batch_evaluate(
            input_path,
            output_dir,
            cases_input,
            args.max_concurrency,
            opts,
            suffix=args.suffix,
        )
    )


if __name__ == "__main__":
    main()
