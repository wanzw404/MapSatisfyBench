"""端到端评测流水线：仿真 → 评分 → 报告，一条命令全部串联。

Usage:
    cd <repo-root>
    .venv/bin/python -m app.scripts.run_pipeline sandbox_test.xlsx [options]

三个阶段通过中间 CSV 路径自动串联：
  Stage 1  多轮对话仿真  → data/outputs/simulator_res/dialogue_*.csv
  Stage 2  批量评分      → data/outputs/evaluation_res/evaluation_result_*.csv
  Stage 3  报告生成      → data/outputs/report/single/<ts>/summary.*

支持 --skip-simulate / --skip-evaluate 从中间阶段恢复。
"""

import argparse
import asyncio
import logging
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.config import settings  # noqa: E402
from app.services.dialogue_recorder import (  # noqa: E402
    DialogueResultWriter,
    create_output_file,
    dialogue_simulator_single,
)
from app.services.excel_parser import read_dialogue_cases  # noqa: E402
from app.scripts.batch_evaluate_from_simulator import batch_evaluate  # noqa: E402
from app.services.eval_summary_service import summarize  # noqa: E402
from app.services.eval_compare_service import compare_grouped  # noqa: E402

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(__name__)

INPUT_DIR = PROJECT_ROOT / "data" / "inputs"
SIMULATOR_RES_DIR = PROJECT_ROOT / "data" / "outputs" / "simulator_res"
EVALUATION_RES_DIR = PROJECT_ROOT / "data" / "outputs" / "evaluation_res"
REPORT_DIR = PROJECT_ROOT / "data" / "outputs" / "report"


def _create_output_file_in(output_dir: Path, input_filename: str, suffix: str | None = None) -> Path:
    """在指定目录下创建仿真 CSV 并写入表头（与 create_output_file 同逻辑，仅目录可控）。"""
    import csv as _csv
    from datetime import datetime as _dt
    from app.services.dialogue_recorder import OUTPUT_HEADERS, _sanitize_filename_suffix
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(input_filename).stem
    suffix_part = ""
    if suffix:
        cleaned = _sanitize_filename_suffix(suffix)
        if cleaned:
            suffix_part = f"_{cleaned}"
    output_name = f"dialogue_{stem}_{timestamp}{suffix_part}.csv"
    output_path = output_dir / output_name
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = _csv.writer(f, quoting=_csv.QUOTE_MINIMAL)
        writer.writerow(OUTPUT_HEADERS)
    return output_path


async def stage_simulate(
    filename: str,
    *,
    model: str,
    suffix: str,
    max_turns: int,
    concurrency: int,
    sandbox: bool,
    streaming: bool,
    thinking: bool,
    output_dir: Path | None = None,
    input_dir: Path | None = None,
) -> Path:
    """Stage 1: 读取用例 → 并发多轮对话仿真 → 写 CSV，返回 CSV 路径。

    Args:
        output_dir: 自定义仿真 CSV 输出目录（默认走 SIMULATOR_RES_DIR）。
        input_dir: 自定义输入文件所在目录（默认走 excel_parser 的 INPUT_DIR）。
    """
    if input_dir:
        from app.schemas.dialogue_simulator import DialogueCase
        from openpyxl import load_workbook
        xlsx_path = input_dir / filename
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        cases = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            rd = dict(zip(headers, row))
            cases.append(DialogueCase(
                task_id=str(rd.get("task_id", "") or "").strip() or None,
                query=str(rd.get("query", "") or ""),
                context=str(rd.get("context", "") or "") or None,
                time=str(rd.get("time", rd.get("current_time", "")) or "") or None,
                location=str(rd.get("location", rd.get("current_location", "")) or "") or None,
                tool=str(rd.get("tool", "") or "") or None,
                persona=str(rd.get("persona", "") or "") or None,
                full_intent=str(rd.get("full_intent", "") or "") or None,
                ground_truth=str(rd.get("ground_truth", "") or "") or None,
            ))
    else:
        cases = read_dialogue_cases(filename)
    if not cases:
        raise SystemExit("Excel 中未找到有效用例")

    mode_str = "沙箱" if sandbox else "真实 API"
    logger.info(
        "Stage 1 — 读取 %d 条用例，%s 模式，并发=%d",
        len(cases), mode_str, concurrency,
    )

    if output_dir:
        output_dir.mkdir(parents=True, exist_ok=True)
        sim_csv = _create_output_file_in(output_dir, filename, suffix)
    else:
        sim_csv = create_output_file(filename, suffix=suffix)
    sem = asyncio.Semaphore(max(1, concurrency))

    async def _run_one(idx: int, case):
        async with sem:
            preview = case.query[:50] + ("..." if len(case.query) > 50 else "")
            logger.info("[%d/%d] 仿真 | %s", idx + 1, len(cases), preview)
            return await dialogue_simulator_single(
                case,
                writer,
                is_sandbox=sandbox,
                max_turns=max_turns,
                streaming=streaming,
                model=model,
                thinking=thinking,
            )

    async with DialogueResultWriter(sim_csv) as writer:
        results = await asyncio.gather(
            *[_run_one(i, c) for i, c in enumerate(cases)]
        )

    n_natural = sum(1 for r in results if r.is_natural_stop)
    logger.info(
        "Stage 1 完成: %d cases, %d 自然终止, 输出: %s",
        len(results), n_natural, sim_csv,
    )
    return sim_csv


async def stage_evaluate(
    sim_csv: Path,
    cases_input: Path,
    *,
    suffix: str,
    eval_concurrency: int,
    language: str,
    enable_verification: bool,
    enable_meta_judge: bool,
    output_dir: Path | None = None,
) -> Path:
    """Stage 2: 读取仿真 CSV → 并发 JudgeAgent 评分 → 返回评分 CSV 路径。

    Args:
        output_dir: 自定义评分 CSV 输出目录（默认走 EVALUATION_RES_DIR）。
    """
    logger.info("Stage 2 — 评分，仿真输入: %s", sim_csv)

    effective_dir = output_dir or EVALUATION_RES_DIR
    effective_dir.mkdir(parents=True, exist_ok=True)
    opts = {
        "language": language,
        "enable_verification": enable_verification,
        "enable_meta_judge": enable_meta_judge,
        "model": "",
    }
    eval_csv = await batch_evaluate(
        input_path=sim_csv,
        output_dir=effective_dir,
        cases_input=cases_input if cases_input.exists() else None,
        max_concurrency=eval_concurrency,
        opts=opts,
        suffix=suffix,
    )
    logger.info("Stage 2 完成，评分结果: %s", eval_csv)
    return eval_csv


def stage_report(
    eval_csv: Path,
    *,
    top_n: int,
    visualize: str,
) -> Path:
    """Stage 3: 读取评分 CSV → 生成汇总报告 → 返回报告路径。"""
    logger.info("Stage 3 — 报告生成，评分输入: %s", eval_csv)

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    products = summarize(
        xlsx_path=eval_csv,
        output_dir=REPORT_DIR,
        top_n=top_n,
        visualize=visualize,
    )
    report_path = products.html_path or products.txt_path
    logger.info("Stage 3 完成，报告: %s", report_path)
    return report_path


async def run_single_model_pipeline(args: argparse.Namespace) -> None:
    """单模型流水线：仿真 → 评分 → 单模型报告。"""
    effective_model = args.model or settings.MODEL_NAME
    effective_suffix = args.suffix if args.suffix is not None else effective_model
    input_xlsx = INPUT_DIR / args.filename

    # ── Stage 1 ──
    if args.skip_simulate:
        if not args.sim_result:
            raise SystemExit("--skip-simulate 需要搭配 --sim-result 指定仿真 CSV 路径")
        sim_csv = Path(args.sim_result)
        if not sim_csv.exists():
            raise SystemExit(f"仿真 CSV 不存在: {sim_csv}")
        logger.info("跳过 Stage 1（仿真），使用已有文件: %s", sim_csv)
    else:
        sim_csv = await stage_simulate(
            args.filename,
            model=effective_model,
            suffix=effective_suffix,
            max_turns=args.max_turns,
            concurrency=args.concurrency,
            sandbox=args.sandbox,
            streaming=args.streaming,
            thinking=args.thinking,
        )

    # ── Stage 2 ──
    if args.skip_evaluate:
        if not args.eval_result:
            raise SystemExit("--skip-evaluate 需要搭配 --eval-result 指定评分 CSV 路径")
        eval_csv = Path(args.eval_result)
        if not eval_csv.exists():
            raise SystemExit(f"评分 CSV 不存在: {eval_csv}")
        logger.info("跳过 Stage 2（评分），使用已有文件: %s", eval_csv)
    else:
        eval_csv = await stage_evaluate(
            sim_csv,
            cases_input=input_xlsx,
            suffix=effective_suffix,
            eval_concurrency=args.eval_concurrency,
            language=args.language,
            enable_verification=args.enable_verification,
            enable_meta_judge=args.enable_meta_judge,
        )

    # ── Stage 3 ──
    report_path = stage_report(
        eval_csv,
        top_n=args.top_n,
        visualize=args.visualize,
    )

    # ── 汇总 ──
    print("\n" + "=" * 60)
    print("Pipeline 完成（单模型）")
    print(f"  仿真结果: {sim_csv}")
    print(f"  评分结果: {eval_csv}")
    print(f"  评估报告: {report_path}")
    print("=" * 60)


async def run_multi_model_pipeline(args: argparse.Namespace) -> None:
    """多模型对比流水线：串行执行每个模型的仿真+评分，最后生成对比报告。"""
    models = [m.strip() for m in args.models.split(",") if m.strip()]
    if len(models) < 2:
        raise SystemExit("--models 至少需要 2 个模型（逗号分隔）")

    input_xlsx = INPUT_DIR / args.filename
    eval_csvs: dict[str, list[Path]] = {}

    for i, model in enumerate(models, 1):
        logger.info(
            "\n%s 模型 [%d/%d]: %s %s",
            "=" * 20, i, len(models), model, "=" * 20,
        )

        sim_csv = await stage_simulate(
            args.filename,
            model=model,
            suffix=model,
            max_turns=args.max_turns,
            concurrency=args.concurrency,
            sandbox=args.sandbox,
            streaming=args.streaming,
            thinking=args.thinking,
        )

        eval_csv = await stage_evaluate(
            sim_csv,
            cases_input=input_xlsx,
            suffix=model,
            eval_concurrency=args.eval_concurrency,
            language=args.language,
            enable_verification=args.enable_verification,
            enable_meta_judge=args.enable_meta_judge,
        )

        eval_csvs.setdefault(model, []).append(eval_csv)
        logger.info("模型 %s 完成 → %s", model, eval_csv)

    # ── 对比报告 ──
    logger.info("\n%s 生成对比报告 %s", "=" * 20, "=" * 20)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    products = compare_grouped(eval_csvs, REPORT_DIR)

    # ── 汇总 ──
    print("\n" + "=" * 60)
    print("Pipeline 完成（多模型对比）")
    print(f"  模型: {', '.join(models)}")
    for model, paths in eval_csvs.items():
        print(f"  [{model}] 评分结果: {paths[0]}")
    print(f"  对比报告: {products.md_path}")
    print(f"  对比数据: {products.json_path}")
    print("=" * 60)


async def run_pipeline(args: argparse.Namespace) -> None:
    if not settings.AI_STUDIO_TOKEN:
        logger.error(
            "AI_STUDIO_TOKEN 未配置，请在 .env 或环境变量中设置后重试。"
        )
        sys.exit(1)

    if args.models:
        return await run_multi_model_pipeline(args)
    return await run_single_model_pipeline(args)


def main():
    parser = argparse.ArgumentParser(
        description="端到端评测流水线：仿真 → 评分 → 报告"
    )
    parser.add_argument(
        "filename",
        help="data/inputs/ 下的 xlsx 文件名",
    )

    model_mutex = parser.add_mutually_exclusive_group()
    model_mutex.add_argument("--model", default="", help="单模型模式：被测 agent 模型名（默认 settings.MODEL_NAME）")
    model_mutex.add_argument(
        "--models", default="",
        help="多模型对比模式：逗号分隔的模型名（如 qwen3-plus,gpt-4o）",
    )

    sim_group = parser.add_argument_group("仿真参数 (Stage 1)")
    sim_group.add_argument("--max-turns", type=int, default=20, help="单 case 最大对话轮次")
    sim_group.add_argument("--concurrency", type=int, default=4, help="仿真并发数")
    sim_group.add_argument("--sandbox", action="store_true", help="沙箱模式（工具走 mock）")
    sim_group.add_argument("--streaming", action="store_true", help="流式 LLM 调用")
    sim_group.add_argument("--thinking", action="store_true", help="推理模式")
    sim_group.add_argument("--suffix", default=None, help="输出文件名后缀（默认用 model 名）")

    eval_group = parser.add_argument_group("评分参数 (Stage 2)")
    eval_group.add_argument("--eval-concurrency", type=int, default=2, help="评分并发数")
    eval_group.add_argument(
        "--language", choices=["chinese", "english"], default="chinese",
        help="评分语言",
    )
    eval_group.add_argument("--enable-verification", action="store_true", help="启用 Web 事实校验")
    eval_group.add_argument("--enable-meta-judge", action="store_true", help="启用 Meta-Judge 审计")

    report_group = parser.add_argument_group("报告参数 (Stage 3)")
    report_group.add_argument("--top-n", type=int, default=5, help="outlier top-N")
    report_group.add_argument(
        "--visualize", choices=["xlsx", "png", "both", "none"], default="both",
        help="报告可视化模式",
    )

    skip_group = parser.add_argument_group("阶段跳过（从中间恢复）")
    skip_group.add_argument("--skip-simulate", action="store_true", help="跳过仿真阶段")
    skip_group.add_argument("--sim-result", default="", help="跳过仿真时指定仿真 CSV 路径")
    skip_group.add_argument("--skip-evaluate", action="store_true", help="跳过评分阶段")
    skip_group.add_argument("--eval-result", default="", help="跳过评分时指定评分 CSV 路径")

    args = parser.parse_args()
    asyncio.run(run_pipeline(args))


if __name__ == "__main__":
    main()
